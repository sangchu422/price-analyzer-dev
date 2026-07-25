from __future__ import annotations

import hashlib
import json
import ntpath
import os
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.cleansing.models import CleanDecision, CleanStatus
from app.cleansing.service import apply_rules
from app.db.base import Base
from app.db.sqlite import configure_sqlite
from app.documents.models import SourceVariant
from app.ingestion.corpus import (
    ingest_corpus,
    preflight_corpus,
    prepare_source_groups,
    scan_supported_files,
)
from app.ingestion.service import ingest_path
from app.quotes.models import RawQuoteItem


def _write_quote(
    path: Path,
    *,
    item_name: str = "BEARING",
    unit_price: int = 2400,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "내역"
    sheet.append(["품명", "규격", "단위", "수량", "단가", "금액"])
    sheet.append([item_name, "6204", "EA", 2, unit_price, unit_price * 2])
    workbook.save(path)


@pytest.fixture
def session() -> Session:
    engine = configure_sqlite(create_engine("sqlite:///:memory:"))
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as database_session:
        yield database_session


def test_preflight_counts_without_reading_source_content(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    root = tmp_path / "quotes"
    _write_quote(root / "A" / "motor.xlsx")
    _write_quote(root / "A" / "motor_보안해제.xlsx")
    (root / "legacy.xls").write_bytes(b"metadata-only")
    (root / "catalog.pdf").write_bytes(b"metadata-only")
    (root / "ignore.csv").write_text("not supported", encoding="utf-8")
    (root / "~$lock.xlsx").write_bytes(b"not a workbook")

    def forbid_content_reads(*args, **kwargs):
        raise AssertionError("preflight must not read source content")

    monkeypatch.setattr(Path, "open", forbid_content_reads)
    monkeypatch.setattr(Path, "read_bytes", forbid_content_reads)

    report = preflight_corpus(root)

    assert report.physical_files == 4
    assert report.files_by_extension == {".pdf": 1, ".xls": 1, ".xlsx": 2}
    assert report.logical_documents == 3
    assert report.variants == 4
    assert report.paired_documents == 1
    assert report.unlocked_variants == 1
    assert report.unlocked_preferred == 1
    assert report.issues == ()


def test_ingest_registers_original_evidence_but_reads_only_unlocked(
    session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    root = tmp_path / "quotes"
    original = root / "drive.xlsx"
    unlocked = root / "drive_보안해제.xlsx"
    _write_quote(original, item_name="PROTECTED")
    _write_quote(unlocked, item_name="UNLOCKED")

    from app.ingestion import service as ingestion_service

    actual_reader = ingestion_service.read_quote
    parsed_paths: list[Path] = []

    def observe_reader(path: Path):
        parsed_paths.append(path)
        return actual_reader(path)

    monkeypatch.setattr(ingestion_service, "read_quote", observe_reader)

    report = ingest_corpus(session, root)

    assert report.documents_ingested == 1
    assert report.documents_unchanged == 0
    assert report.documents_failed == 0
    assert parsed_paths == [unlocked.resolve()]
    assert session.scalar(select(func.count(SourceVariant.id))) == 2
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1
    assert session.scalar(select(RawQuoteItem)).item_name_raw == "UNLOCKED"


def test_ingest_is_idempotent_and_applies_base_and_outlier_rules(
    session: Session,
    tmp_path: Path,
) -> None:
    root = tmp_path / "quotes"
    for index, price in enumerate((1000, 1000, 100000), start=1):
        _write_quote(
            root / f"quote-{index}.xlsx",
            item_name="BEARING",
            unit_price=price,
        )

    first = ingest_corpus(session, root)
    first_rows = session.scalar(select(func.count(RawQuoteItem.id)))
    first_decisions = session.scalar(select(func.count(CleanDecision.id)))
    second = ingest_corpus(session, root)

    assert first.raw_items_created == 3
    assert first.base_decisions_created == 3
    assert first.outlier_decisions_created == 1
    assert first.latest_status_counts == {
        CleanStatus.INCLUDED.value: 2,
        CleanStatus.EXCLUDED.value: 0,
        CleanStatus.REVIEW_REQUIRED.value: 1,
    }
    assert second.documents_ingested == 0
    assert second.documents_unchanged == 3
    assert second.raw_items_created == 0
    assert second.base_decisions_created == 0
    assert second.outlier_decisions_created == 0
    assert session.scalar(select(func.count(RawQuoteItem.id))) == first_rows
    assert session.scalar(select(func.count(CleanDecision.id))) == first_decisions


def test_failed_document_does_not_abort_good_documents_and_is_accounted(
    session: Session,
    tmp_path: Path,
) -> None:
    root = tmp_path / "quotes"
    _write_quote(root / "good.xlsx", item_name="GOOD")
    workbook = Workbook()
    workbook.active.append(["unknown", "layout"])
    workbook.save(root / "unsupported.xlsx")

    report = ingest_corpus(session, root)

    assert report.documents_ingested == 1
    assert report.documents_failed == 1
    assert [result.status for result in report.documents] == [
        "INGESTED",
        "FAILED",
    ]
    failure = report.failures[0]
    assert failure.logical_name == "unsupported"
    assert failure.error_code == "UNSUPPORTED_LAYOUT"
    assert "\\" not in failure.logical_name
    assert str(tmp_path) not in json.dumps(report.to_dict(), ensure_ascii=False)
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1


def test_failed_group_report_preserves_relative_variant_hash_evidence(
    session: Session,
    tmp_path: Path,
) -> None:
    root = tmp_path / "quotes"
    original = root / "unsupported.xlsx"
    unlocked = root / "unsupported_보안해제.xlsx"
    _write_quote(original, item_name="PROTECTED")
    workbook = Workbook()
    workbook.active.append(["unknown", "layout"])
    workbook.save(unlocked)

    report = ingest_corpus(session, root)

    failure = report.failures[0].to_dict()
    assert failure["preferred_path"] == "unsupported_보안해제.xlsx"
    assert failure["preferred_sha256"] == hashlib.sha256(
        unlocked.read_bytes()
    ).hexdigest()
    assert failure["variants"] == [
        {
            "path": "unsupported.xlsx",
            "sha256": hashlib.sha256(original.read_bytes()).hexdigest(),
        },
        {
            "path": "unsupported_보안해제.xlsx",
            "sha256": hashlib.sha256(unlocked.read_bytes()).hexdigest(),
        },
    ]
    assert str(tmp_path) not in json.dumps(
        report.to_dict(),
        ensure_ascii=False,
    )


def test_failed_group_reports_hash_unavailable_without_leaking_path(
    session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    root = tmp_path / "quotes"
    quote = root / "unsupported.xlsx"
    quote.parent.mkdir()
    workbook = Workbook()
    workbook.active.append(["unknown", "layout"])
    workbook.save(quote)

    def fail_audit_hash(path: Path) -> str:
        raise OSError(f"sensitive absolute path: {path}")

    monkeypatch.setattr("app.ingestion.corpus.sha256", fail_audit_hash)

    report = ingest_corpus(session, root)

    failure = report.failures[0].to_dict()
    assert failure["preferred_sha256"] is None
    assert failure["variants"] == [
        {
            "path": "unsupported.xlsx",
            "sha256": None,
            "error_code": "HASH_UNAVAILABLE",
        }
    ]
    assert str(tmp_path) not in json.dumps(
        report.to_dict(),
        ensure_ascii=False,
    )


def test_missing_root_does_not_append_outliers_or_mutate_existing_data(
    session: Session,
    tmp_path: Path,
) -> None:
    root = tmp_path / "quotes"
    for index, price in enumerate((1000, 1000, 100000), start=1):
        quote = root / f"quote-{index}.xlsx"
        _write_quote(quote, unit_price=price)
        variant = ingest_path(session, quote, root=root)
        apply_rules(session, variant.raw_items[0])
    session.commit()
    counts_before = (
        session.scalar(select(func.count(SourceVariant.id))),
        session.scalar(select(func.count(RawQuoteItem.id))),
        session.scalar(select(func.count(CleanDecision.id))),
    )

    report = ingest_corpus(session, tmp_path / "missing")

    assert report.documents_failed == 1
    assert report.failures[0].error_code == "QUOTE_ROOT_NOT_FOUND"
    assert report.documents[0].status == "FAILED"
    assert (
        session.scalar(select(func.count(SourceVariant.id))),
        session.scalar(select(func.count(RawQuoteItem.id))),
        session.scalar(select(func.count(CleanDecision.id))),
    ) == counts_before


def test_missing_root_is_an_explicit_preflight_error(tmp_path: Path) -> None:
    report = preflight_corpus(tmp_path / "missing")

    assert report.root_available is False
    assert report.issues[0].error_code == "QUOTE_ROOT_NOT_FOUND"


def test_quote_root_file_is_reported_as_not_a_directory(
    tmp_path: Path,
) -> None:
    root_file = tmp_path / "quote-root.xlsx"
    _write_quote(root_file)

    report = preflight_corpus(root_file)

    assert report.root_available is False
    assert report.issues[0].error_code == "QUOTE_ROOT_NOT_DIRECTORY"


@pytest.mark.real_corpus
@pytest.mark.skipif(
    not os.environ.get("PRICE_ANALYZER_REAL_QUOTE_ROOT"),
    reason="set PRICE_ANALYZER_REAL_QUOTE_ROOT for read-only local audit",
)
def test_real_corpus_preflight_is_read_only_and_records_current_observation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    root = Path(os.environ["PRICE_ANALYZER_REAL_QUOTE_ROOT"])

    def forbid_content_reads(*args, **kwargs):
        raise AssertionError("real corpus preflight must not read source content")

    monkeypatch.setattr(Path, "open", forbid_content_reads)
    monkeypatch.setattr(Path, "read_bytes", forbid_content_reads)

    report = preflight_corpus(root)
    paths = scan_supported_files(root)
    groups, issues = prepare_source_groups(paths, root)
    paired_groups = [
        group
        for group in groups
        if any(_is_unlocked_variant(path) for path in group.variants)
        and any(not _is_unlocked_variant(path) for path in group.variants)
    ]

    assert report.root_available
    assert report.physical_files == 48
    assert report.logical_documents == 36
    assert report.unlocked_preferred == 15
    assert report.paired_documents == 12
    assert len(paths) == report.physical_files
    assert issues == []
    assert len(paired_groups) == 12
    assert all(
        _is_unlocked_variant(group.preferred)
        for group in paired_groups
    )


def _is_unlocked_variant(path: Path) -> bool:
    stem = ntpath.normcase(path.stem.strip())
    suffix = ntpath.normcase("_보안해제")
    return stem.endswith(suffix)
