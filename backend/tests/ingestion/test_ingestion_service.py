from __future__ import annotations

import json
import os
import zlib
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, func, inspect as sa_inspect, select
from sqlalchemy.orm import Session

from app.db.base import Base
from app.db.models import RawQuoteItem, SourceDocument, SourceVariant
from app.db.sqlite import configure_sqlite
from app.ingestion.readers import (
    ParsedRow,
    UnsafeQuoteFileError,
    read_quote,
    read_xlsx,
)
from app.ingestion.service import (
    SourceFileChangedError,
    UnsupportedQuoteLayoutError,
    ingest_group,
    ingest_path,
    parsing_variant_for,
    preferred_variant_for,
)
from app.ingestion.source_selector import build_source_groups


@pytest.fixture
def session() -> Session:
    engine = configure_sqlite(create_engine("sqlite:///:memory:"))
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as database_session:
        yield database_session


def _write_quote(
    path: Path,
    *,
    item_name: str = "SERVO MOTOR",
    unit_price: object = 500000,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "단위설비1"
    sheet.append(["견적서"])
    for _ in range(5):
        sheet.append([])
    sheet.append(["품명", "규격", "단위", "수량", "단가", "금액", "메이커"])
    sheet.append(
        [item_name, "AC 220V", "EA", "2", unit_price, "1000000", "ACME"]
    )
    workbook.save(path)


def _write_layout_fixture(
    path: Path,
    fixture_name: str,
) -> None:
    fixtures_path = (
        Path(__file__).resolve().parents[1]
        / "fixtures"
        / "parser_layouts.json"
    )
    fixture = json.loads(fixtures_path.read_text(encoding="utf-8"))[
        fixture_name
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = fixture["sheet"]
    for row in fixture["rows"]:
        sheet.append(row)
    workbook.save(path)


@pytest.mark.parametrize(
    ("fixture_name", "expected_item"),
    [
        ("standard", "SERVO MOTOR"),
        ("assembly_device", "ROBOT"),
        ("assembly_content", "SAFETY FENCE"),
        ("assembly_fallback", "CONTROL PANEL"),
    ],
)
def test_verified_legacy_layouts_do_not_silently_return_zero_rows(
    tmp_path: Path,
    fixture_name: str,
    expected_item: str,
) -> None:
    quote = tmp_path / f"{fixture_name}.xlsx"
    _write_layout_fixture(quote, fixture_name)

    rows = read_quote(quote)

    assert len(rows) == 1
    assert rows[0].item_name == expected_item
    assert rows[0].sheet is not None
    assert rows[0].row == 3


def test_headerless_fixed_column_fallback_is_auditable(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "headerless.xlsx"
    _write_layout_fixture(quote, "headerless_fixed_columns")

    rows = read_quote(quote)

    assert [row.item_name for row in rows] == ["PHOTO SENSOR", "BEARING"]
    assert rows[0].quantity == "2"
    assert rows[0].unit == "EA"
    assert rows[0].unit_price == "11100"
    assert rows[0].row == 2
    assert rows[0].cells == "C2:H2"
    assert rows[0].warnings == ("FALLBACK_FIXED_C_E_F_H",)
    variant = ingest_path(session, quote, root=tmp_path)
    assert json.loads(variant.raw_items[0].parse_warnings_json) == [
        "FALLBACK_FIXED_C_E_F_H"
    ]


def test_ingestion_preserves_exact_variant_and_cell_provenance(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "설비 견적_보안해제.xlsx"
    _write_quote(quote)

    variant = ingest_path(session, quote, root=tmp_path)
    item = session.scalar(select(RawQuoteItem))

    assert item is not None
    assert variant.path == "설비 견적_보안해제.xlsx"
    assert item.source_variant is variant
    assert item.source_sheet == "단위설비1"
    assert item.source_page is None
    assert item.source_row == 8
    assert item.source_cells == "A8:G8"
    assert item.item_name_raw == "SERVO MOTOR"
    assert item.spec_raw == "AC 220V"
    assert item.unit_raw == "EA"
    assert item.quantity_raw == "2"
    assert item.unit_price_raw == "500000"
    assert item.amount_raw == "1000000"
    assert item.maker_raw == "ACME"


def test_same_file_content_is_idempotent(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "설비 견적.xlsx"
    _write_quote(quote)

    first = ingest_path(session, quote, root=tmp_path)
    second = ingest_path(session, quote, root=tmp_path)

    assert second.id == first.id
    assert session.scalar(select(func.count(SourceVariant.id))) == 1
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1


def test_changed_content_at_an_existing_path_is_rejected(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "설비 견적.xlsx"
    _write_quote(quote, item_name="FIRST")
    ingest_path(session, quote, root=tmp_path)
    _write_quote(quote, item_name="CHANGED")

    with pytest.raises(ValueError, match="content changed at immutable source path"):
        ingest_path(session, quote, root=tmp_path)

    assert session.scalar(select(func.count(SourceVariant.id))) == 1
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1


@pytest.mark.skipif(os.name != "nt", reason="Windows path identity")
def test_existing_path_identity_is_case_insensitive_on_windows(
    session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    quote_folder = tmp_path / "Quotes"
    quote_folder.mkdir()
    quote = quote_folder / "Quote.xlsx"
    _write_quote(quote, item_name="FIRST")
    monkeypatch.chdir(tmp_path)
    ingest_path(session, Path("Quotes/Quote.xlsx"))
    _write_quote(quote, item_name="CHANGED")

    with pytest.raises(ValueError, match="content changed at immutable source path"):
        ingest_path(session, Path("QUOTES/QUOTE.xlsx"))

    assert session.scalar(select(func.count(SourceVariant.id))) == 1


def test_duplicate_content_at_different_paths_preserves_both_evidence_records(
    session: Session,
    tmp_path: Path,
) -> None:
    first = tmp_path / "첫 견적.xlsx"
    second = tmp_path / "다른 견적.xlsx"
    _write_quote(first)
    second.write_bytes(first.read_bytes())
    ingest_path(session, first, root=tmp_path)
    ingest_path(session, second, root=tmp_path)

    assert session.scalar(select(func.count(SourceDocument.id))) == 2
    assert session.scalar(select(func.count(SourceVariant.id))) == 2
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 2


def test_group_ingestion_registers_both_variants_but_parses_only_unlocked(
    session: Session,
    tmp_path: Path,
) -> None:
    original = tmp_path / "설비 견적.xlsx"
    unlocked = tmp_path / "설비 견적_보안해제.xlsx"
    _write_quote(original, item_name="LOCKED COPY")
    _write_quote(unlocked, item_name="UNLOCKED COPY")
    group = build_source_groups([original, unlocked], root=tmp_path)[0]

    preferred = ingest_group(session, group, root=tmp_path)

    variants = session.scalars(
        select(SourceVariant).order_by(SourceVariant.path)
    ).all()
    rows = session.scalars(select(RawQuoteItem)).all()
    assert preferred.path == "설비 견적_보안해제.xlsx"
    assert len(variants) == 2
    assert sum(
        variant.selected_for_parsing_at_ingest
        for variant in variants
    ) == 1
    assert len(rows) == 1
    assert rows[0].item_name_raw == "UNLOCKED COPY"
    assert rows[0].source_variant.path == "설비 견적_보안해제.xlsx"


def test_identical_group_content_is_stored_and_parsed_only_once(
    session: Session,
    tmp_path: Path,
) -> None:
    original = tmp_path / "동일 견적.xlsx"
    unlocked = tmp_path / "동일 견적_보안해제.xlsx"
    _write_quote(unlocked, item_name="SAME CONTENT")
    original.write_bytes(unlocked.read_bytes())
    group = build_source_groups([original, unlocked], root=tmp_path)[0]

    first = ingest_group(session, group, root=tmp_path)
    second = ingest_group(session, group, root=tmp_path)

    assert first.id == second.id
    assert first.path == "동일 견적_보안해제.xlsx"
    assert session.scalar(select(func.count(SourceVariant.id))) == 2
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1
    assert {variant.path for variant in first.document.variants} == {
        "동일 견적.xlsx",
        "동일 견적_보안해제.xlsx",
    }


def test_byte_identical_locked_then_unlocked_retains_unlocked_precedence(
    session: Session,
    tmp_path: Path,
) -> None:
    original = tmp_path / "동일 순차 견적.xlsx"
    unlocked = tmp_path / "동일 순차 견적_보안해제.xlsx"
    _write_quote(original, item_name="SHARED ROW")
    unlocked.write_bytes(original.read_bytes())

    original_variant = ingest_path(session, original, root=tmp_path)
    unlocked_variant = ingest_path(session, unlocked, root=tmp_path)

    assert original_variant.id != unlocked_variant.id
    assert original_variant.path == "동일 순차 견적.xlsx"
    assert unlocked_variant.path == "동일 순차 견적_보안해제.xlsx"
    assert original_variant.security_state == "UNKNOWN"
    assert unlocked_variant.security_state == "UNLOCKED"
    assert not original_variant.selected_for_parsing_at_ingest
    assert unlocked_variant.selected_for_parsing_at_ingest
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 1
    assert preferred_variant_for(unlocked_variant.document) is unlocked_variant
    assert parsing_variant_for(session, unlocked_variant) is original_variant


def test_locked_then_unlocked_keeps_evidence_and_selects_unlocked(
    session: Session,
    tmp_path: Path,
) -> None:
    original = tmp_path / "설비 견적.xlsx"
    unlocked = tmp_path / "설비 견적_보안해제.xlsx"
    _write_quote(original, item_name="ORIGINAL EVIDENCE")
    _write_quote(unlocked, item_name="UNLOCKED EVIDENCE")

    original_variant = ingest_path(session, original, root=tmp_path)
    unlocked_variant = ingest_path(session, unlocked, root=tmp_path)

    document = session.scalar(select(SourceDocument))
    assert document is not None
    assert original_variant.document_id == unlocked_variant.document_id
    assert not original_variant.selected_for_parsing_at_ingest
    assert unlocked_variant.selected_for_parsing_at_ingest
    assert {row.item_name_raw for row in document.raw_items} == {
        "ORIGINAL EVIDENCE",
        "UNLOCKED EVIDENCE",
    }


def test_absolute_ingestion_requires_a_stable_containing_root(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "설비 견적.xlsx"
    _write_quote(quote)

    with pytest.raises(ValueError, match="explicit stable root"):
        ingest_path(session, quote)
    with pytest.raises(ValueError, match="outside the declared root"):
        ingest_path(session, quote, root=tmp_path / "other")

    assert session.scalar(select(func.count(SourceVariant.id))) == 0


def test_parse_errors_leave_no_partial_database_rows(
    session: Session,
    tmp_path: Path,
) -> None:
    corrupt = tmp_path / "손상 견적.xlsx"
    corrupt.write_bytes(b"not an xlsx archive")

    with pytest.raises(Exception):
        ingest_path(session, corrupt, root=tmp_path)

    assert session.scalar(select(func.count(SourceDocument.id))) == 0
    assert session.scalar(select(func.count(SourceVariant.id))) == 0
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 0


def test_supported_file_with_unknown_layout_rolls_back_visibly(
    session: Session,
    tmp_path: Path,
) -> None:
    quote = tmp_path / "unknown-layout.xlsx"
    workbook = Workbook()
    workbook.active.append(["알 수 없는", "양식"])
    workbook.save(quote)

    with pytest.raises(
        UnsupportedQuoteLayoutError,
        match="no quote rows matched",
    ):
        ingest_path(session, quote, root=tmp_path)

    assert session.scalar(select(func.count(SourceDocument.id))) == 0
    assert session.scalar(select(func.count(SourceVariant.id))) == 0


def test_file_swap_between_hash_and_parse_rolls_back_all_ingestion(
    session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    quote = tmp_path / "changing.xlsx"
    _write_quote(quote, item_name="BEFORE")
    from app.ingestion import service as ingestion_service

    actual_reader = ingestion_service.read_quote

    def read_then_replace(path: Path) -> list[ParsedRow]:
        rows = actual_reader(path)
        _write_quote(path, item_name="AFTER")
        return rows

    monkeypatch.setattr(ingestion_service, "read_quote", read_then_replace)

    with pytest.raises(SourceFileChangedError, match="changed while parsing"):
        ingest_path(session, quote, root=tmp_path)

    assert session.scalar(select(func.count(SourceDocument.id))) == 0
    assert session.scalar(select(func.count(SourceVariant.id))) == 0
    assert session.scalar(select(func.count(RawQuoteItem.id))) == 0


def test_ingestion_success_does_not_commit_unrelated_outer_work(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "transaction-success.sqlite3"
    engine = configure_sqlite(
        create_engine(f"sqlite:///{database_path.as_posix()}")
    )
    Base.metadata.create_all(engine)
    quote = tmp_path / "quote.xlsx"
    _write_quote(quote)
    with Session(engine, expire_on_commit=False) as session:
        unrelated = SourceDocument(logical_name="unrelated-pending")
        session.add(unrelated)

        ingest_path(session, quote, root=tmp_path)

        assert session.in_transaction()
        assert sa_inspect(unrelated).persistent
        with Session(engine) as observer:
            assert observer.scalar(
                select(func.count(SourceDocument.id))
            ) == 0
        session.commit()
    with Session(engine) as observer:
        assert observer.scalar(select(func.count(SourceDocument.id))) == 2


def test_ingestion_failure_rolls_back_only_its_savepoint(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "transaction-failure.sqlite3"
    engine = configure_sqlite(
        create_engine(f"sqlite:///{database_path.as_posix()}")
    )
    Base.metadata.create_all(engine)
    quote = tmp_path / "unknown.xlsx"
    workbook = Workbook()
    workbook.active.append(["unknown"])
    workbook.save(quote)
    with Session(engine, expire_on_commit=False) as session:
        unrelated = SourceDocument(logical_name="keep-me-pending")
        session.add(unrelated)

        with pytest.raises(UnsupportedQuoteLayoutError):
            ingest_path(session, quote, root=tmp_path)

        assert session.in_transaction()
        assert sa_inspect(unrelated).persistent
        assert session.scalar(select(func.count(SourceDocument.id))) == 1
        assert session.scalar(select(func.count(SourceVariant.id))) == 0
        with Session(engine) as observer:
            assert observer.scalar(
                select(func.count(SourceDocument.id))
            ) == 0
        session.commit()
    with Session(engine) as observer:
        assert observer.scalar(select(func.count(SourceDocument.id))) == 1


def test_clean_caller_rollback_removes_ingestion_after_savepoint(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "clean-rollback.sqlite3"
    engine = create_engine(f"sqlite:///{database_path.as_posix()}")
    configure_sqlite(engine)
    configure_sqlite(engine)
    Base.metadata.create_all(engine)
    quote = tmp_path / "rollback-quote.xlsx"
    _write_quote(quote)
    with engine.connect() as connection:
        assert connection.connection.driver_connection.isolation_level is None
        assert connection.exec_driver_sql(
            "PRAGMA foreign_keys"
        ).scalar_one() == 1

    with Session(engine) as session:
        ingest_path(session, quote, root=tmp_path)
        session.rollback()

    with Session(engine) as observer:
        assert observer.scalar(select(func.count(SourceDocument.id))) == 0
        assert observer.scalar(select(func.count(SourceVariant.id))) == 0
        assert observer.scalar(select(func.count(RawQuoteItem.id))) == 0


def test_clean_caller_commit_persists_ingestion_after_savepoint(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "clean-commit.sqlite3"
    engine = configure_sqlite(
        create_engine(f"sqlite:///{database_path.as_posix()}")
    )
    Base.metadata.create_all(engine)
    quote = tmp_path / "commit-quote.xlsx"
    _write_quote(quote)

    with Session(engine) as session:
        ingest_path(session, quote, root=tmp_path)
        session.commit()

    with Session(engine) as observer:
        assert observer.scalar(select(func.count(SourceDocument.id))) == 1
        assert observer.scalar(select(func.count(SourceVariant.id))) == 1
        assert observer.scalar(select(func.count(RawQuoteItem.id))) == 1


def test_unsupported_extension_is_rejected_without_database_writes(
    session: Session,
    tmp_path: Path,
) -> None:
    unsupported = tmp_path / "견적서.csv"
    unsupported.write_text("품명,단가\nMOTOR,1000", encoding="utf-8")

    with pytest.raises(ValueError, match="unsupported quote extension"):
        ingest_path(session, unsupported, root=tmp_path)

    assert session.scalar(select(func.count(SourceDocument.id))) == 0


def test_xls_reader_uses_same_parsed_row_contract(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class FakeSheet:
        name = "내역"
        nrows = 2
        ncols = 3

        def cell_value(self, row: int, column: int) -> object:
            rows = [
                ["품명", "규격", "단가"],
                ["BEARING", "6204", 2400],
            ]
            return rows[row][column]

    class FakeBook:
        def sheets(self) -> list[FakeSheet]:
            return [FakeSheet()]

    monkeypatch.setattr(
        "app.ingestion.readers.xlrd.open_workbook",
        lambda _: FakeBook(),
    )
    quote = tmp_path / "legacy.xls"
    quote.write_bytes(b"fixture")

    rows = read_quote(quote)

    assert rows == [
        ParsedRow(
            sheet="내역",
            page=None,
            row=2,
            cells="A2:C2",
            item_name="BEARING",
            spec="6204",
            unit=None,
            quantity=None,
            unit_price="2400",
            amount=None,
            maker=None,
        )
    ]


def test_pdf_reader_records_page_provenance(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class FakePage:
        def extract_text(self) -> str:
            return "품명  규격  단가\nBEARING  6204  2400"

    class FakePdf:
        pages = [FakePage()]

    monkeypatch.setattr("app.ingestion.readers.PdfReader", lambda _: FakePdf())
    quote = tmp_path / "quote.pdf"
    quote.write_bytes(b"fixture")

    rows = read_quote(quote)

    assert rows[0].sheet is None
    assert rows[0].page == 1
    assert rows[0].row is None
    assert rows[0].cells is None
    assert rows[0].item_name == "BEARING"
    assert rows[0].unit_price == "2400"


def test_pdf_reader_rejects_page_and_extracted_text_limits(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    class FakePage:
        def extract_text(self) -> str:
            return "abcdef"

    class FakePdf:
        pages = [FakePage(), FakePage()]

    monkeypatch.setattr(readers, "PdfReader", lambda _: FakePdf())
    quote = tmp_path / "bounded.pdf"
    quote.write_bytes(b"fixture")
    monkeypatch.setattr(readers, "MAX_PDF_PAGES", 1)
    with pytest.raises(UnsafeQuoteFileError):
        read_quote(quote)

    monkeypatch.setattr(readers, "MAX_PDF_PAGES", 2)
    monkeypatch.setattr(readers, "MAX_PDF_EXTRACTED_TEXT_CHARS", 5)
    with pytest.raises(UnsafeQuoteFileError):
        read_quote(quote)


def test_xlsx_reader_rejects_unknown_dimensions(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    class FakeSheet:
        max_row = None
        max_column = None

    class FakeWorkbook:
        worksheets = [FakeSheet()]

        def close(self) -> None:
            pass

    monkeypatch.setattr(readers, "_validate_xlsx_archive", lambda _: None)
    monkeypatch.setattr(
        readers,
        "load_workbook",
        lambda *args, **kwargs: FakeWorkbook(),
    )

    with pytest.raises(UnsafeQuoteFileError):
        read_xlsx(tmp_path / "unknown-dimensions.xlsx")


def test_pdf_flate_decode_is_bounded_and_other_filters_are_rejected() -> None:
    from pypdf.generic import EncodedStreamObject, NameObject

    from app.ingestion import readers

    compressed = zlib.compress(b"A" * 100)
    assert readers._bounded_flate_size(compressed, 100) == 100
    with pytest.raises(UnsafeQuoteFileError):
        readers._bounded_flate_size(compressed, 99)

    stream = EncodedStreamObject()
    stream._data = b"encoded"
    stream[NameObject("/Filter")] = NameObject("/LZWDecode")

    class FakePage:
        def raw_get(self, name):
            return stream

    with pytest.raises(UnsafeQuoteFileError):
        readers._bounded_pdf_page_content(
            FakePage(),
            decoded_remaining=100,
        )


def test_pdf_rejects_runlength_tounicode_before_decoder_is_invoked(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from pypdf.generic import (
        DictionaryObject,
        EncodedStreamObject,
        NameObject,
    )

    from app.ingestion import readers

    to_unicode = EncodedStreamObject()
    to_unicode._data = b"\x80A\x80B\x80C"
    to_unicode[NameObject("/Filter")] = NameObject("/RunLengthDecode")
    font = DictionaryObject(
        {NameObject("/ToUnicode"): to_unicode}
    )
    resources = DictionaryObject(
        {
            NameObject("/Font"): DictionaryObject(
                {NameObject("/F1"): font}
            )
        }
    )
    decoder_called = False

    class FakePage:
        def raw_get(self, name):
            if name == "/Resources":
                return resources
            raise KeyError(name)

        def extract_text(self) -> str:
            nonlocal decoder_called
            decoder_called = True
            to_unicode.get_data()
            return ""

    class FakePdf:
        pages = [FakePage()]

    monkeypatch.setattr(readers, "PdfReader", lambda _: FakePdf())
    quote = tmp_path / "resource-bomb.pdf"
    quote.write_bytes(b"fixture")

    with pytest.raises(UnsafeQuoteFileError):
        read_quote(quote)

    assert not decoder_called


def test_pdf_allows_bounded_dct_image_that_text_extraction_skips(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from pypdf.generic import (
        DictionaryObject,
        EncodedStreamObject,
        NameObject,
        NumberObject,
    )

    from app.ingestion import readers

    image = EncodedStreamObject()
    image._data = b"bounded-jpeg"
    image[NameObject("/Subtype")] = NameObject("/Image")
    image[NameObject("/Filter")] = NameObject("/DCTDecode")
    image[NameObject("/Width")] = NumberObject(100)
    image[NameObject("/Height")] = NumberObject(80)
    resources = DictionaryObject(
        {
            NameObject("/XObject"): DictionaryObject(
                {NameObject("/Im1"): image}
            )
        }
    )

    class FakePage:
        def raw_get(self, name):
            if name == "/Resources":
                return resources
            raise KeyError(name)

        def extract_text(self) -> str:
            return ""

    class FakePdf:
        pages = [FakePage()]

    monkeypatch.setattr(readers, "PdfReader", lambda _: FakePdf())
    quote = tmp_path / "bounded-image.pdf"
    quote.write_bytes(b"fixture")

    assert read_quote(quote) == []


def test_pdf_raw_preflight_rejects_escaped_runlength_before_reader(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from pypdf import filters

    from app.ingestion import readers

    reader_called = False
    decoder_called = False

    def fail_reader(*args, **kwargs):
        nonlocal reader_called
        reader_called = True
        raise AssertionError("PdfReader must not be constructed")

    def fail_decoder(*args, **kwargs):
        nonlocal decoder_called
        decoder_called = True
        raise AssertionError("RunLength decoder must not run")

    monkeypatch.setattr(readers, "PdfReader", fail_reader)
    monkeypatch.setattr(filters.RunLengthDecode, "decode", fail_decoder)
    attack = tmp_path / "object-stream-attack.pdf"
    attack.write_bytes(
        b"%PDF-1.7\n"
        b"1 0 obj\n"
        b"<< /Type /ObjStm /N 1 /First 4 "
        b"/Filter /Run#4cengthDecode /Length 7 >>\n"
        b"stream\n\x80A\x80B\nendstream\nendobj\n%%EOF\n"
    )

    with pytest.raises(UnsafeQuoteFileError):
        read_quote(attack)

    assert not reader_called
    assert not decoder_called


def test_pdf_raw_preflight_ignores_filter_text_in_comments_and_strings(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    class FakePdf:
        pages = []

    monkeypatch.setattr(readers, "PdfReader", lambda _: FakePdf())
    quote = tmp_path / "lexical-comments.pdf"
    quote.write_bytes(
        b"%PDF-1.7\n"
        b"% /Filter /DCTDecode\n"
        b"1 0 obj << /Note (/Filter /DCTDecode) "
        b"/Filter /FlateDecode >> endobj\n%%EOF\n"
    )

    assert read_quote(quote) == []


def test_pdf_raw_preflight_rejects_filter_chains_and_token_overflow(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    chain = tmp_path / "filter-chain.pdf"
    chain.write_bytes(
        b"%PDF-1.7\n"
        b"1 0 obj << /Filter [/ASCII85Decode /FlateDecode] >> endobj\n"
    )
    with pytest.raises(UnsafeQuoteFileError):
        read_quote(chain)

    monkeypatch.setattr(readers, "MAX_PDF_LEXICAL_TOKENS", 5)
    overflow = tmp_path / "token-overflow.pdf"
    overflow.write_bytes(b"%PDF-1.7\n1 2 3 4 5 6 7\n")
    with pytest.raises(UnsafeQuoteFileError):
        read_quote(overflow)


def test_pdf_raw_preflight_rejects_cr_only_hidden_filter_before_reader(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    reader_called = False

    def fail_reader(*args, **kwargs):
        nonlocal reader_called
        reader_called = True
        raise AssertionError("PdfReader must not be constructed")

    monkeypatch.setattr(readers, "PdfReader", fail_reader)
    attack = tmp_path / "cr-comment-attack.pdf"
    attack.write_bytes(
        b"%PDF-1.7\r"
        b"% harmless comment\r"
        b"1 0 obj << /Filter /LZWDecode >> endobj\r%%EOF\r"
    )

    with pytest.raises(UnsafeQuoteFileError):
        read_quote(attack)
    assert not reader_called


def test_pdf_raw_preflight_does_not_skip_orphan_stream_decoy(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    reader_called = False

    def fail_reader(*args, **kwargs):
        nonlocal reader_called
        reader_called = True
        raise AssertionError("PdfReader must not be constructed")

    monkeypatch.setattr(readers, "PdfReader", fail_reader)
    attack = tmp_path / "orphan-stream-decoy.pdf"
    attack.write_bytes(
        b"%PDF-1.7\nstream\n"
        b"2 0 obj << /Type /ObjStm "
        b"/Filter /Run#4cengthDecode >> endobj "
        b"binary-endstream-decoy endstream\n"
    )

    with pytest.raises(UnsafeQuoteFileError):
        read_quote(attack)
    assert not reader_called


def test_pdf_raw_preflight_rejects_indirect_filter_and_false_boundary(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from app.ingestion import readers

    reader_called = False

    def fail_reader(*args, **kwargs):
        nonlocal reader_called
        reader_called = True
        raise AssertionError("PdfReader must not be constructed")

    monkeypatch.setattr(readers, "PdfReader", fail_reader)
    attack = tmp_path / "indirect-filter.pdf"
    attack.write_bytes(
        b"%PDF-1.7\n"
        b"1 0 obj << /Length 30 >> stream\n"
        b"abc /Filter 9 0 R "
        b"endstream false-boundary endstream\nendobj\n"
    )

    with pytest.raises(UnsafeQuoteFileError):
        read_quote(attack)
    assert not reader_called
