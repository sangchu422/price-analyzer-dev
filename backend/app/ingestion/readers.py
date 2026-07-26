"""Reader-neutral extraction of quote rows with source provenance."""

from __future__ import annotations

import re
import threading
import zipfile
import zlib
from contextlib import contextmanager
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from pypdf.generic import (
    ArrayObject,
    DictionaryObject,
    IndirectObject,
    StreamObject,
)


SUPPORTED_QUOTE_EXTENSIONS = frozenset({".xlsx", ".xls", ".pdf"})
MAX_XLSX_ARCHIVE_ENTRIES = 5_000
MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
MAX_XLSX_WORKSHEETS = 200
MAX_XLSX_WORKSHEET_ROWS = 200_000
MAX_XLSX_WORKSHEET_COLUMNS = 500
MAX_XLSX_WORKSHEET_CELLS = 1_000_000
MAX_XLSX_TOTAL_CELLS = 1_000_000
MAX_XLS_SHEETS = 200
MAX_XLS_SHEET_ROWS = 200_000
MAX_XLS_SHEET_COLUMNS = 500
MAX_XLS_SHEET_CELLS = 2_000_000
MAX_XLS_TOTAL_CELLS = 5_000_000
MAX_PDF_PAGES = 200
MAX_PDF_COMPRESSED_CONTENT_BYTES = 25 * 1024 * 1024
MAX_PDF_DECODED_CONTENT_BYTES = 32 * 1024 * 1024
MAX_PDF_EXTRACTED_TEXT_CHARS = 5_000_000
MAX_PDF_EXTRACTED_ROWS = 200_000
MAX_PDF_TOTAL_FLATE_DECODED_BYTES = 32 * 1024 * 1024
MAX_PDF_REACHABLE_OBJECTS = 50_000
MAX_PDF_RESOURCE_DEPTH = 50
MAX_PDF_IMAGE_COUNT = 1_000
MAX_PDF_IMAGE_RAW_BYTES = 16 * 1024 * 1024
MAX_PDF_IMAGE_PIXELS = 25_000_000
MAX_PDF_TOTAL_IMAGE_PIXELS = 100_000_000
MAX_PDF_RAW_BYTES = 25 * 1024 * 1024
MAX_PDF_LEXICAL_TOKENS = 2_000_000
MAX_PDF_LEXICAL_NAMES = 500_000
MAX_PDF_LEXICAL_OBJECTS = 200_000
MAX_PDF_LEXICAL_REFERENCES = 500_000
MAX_PDF_LEXICAL_DEPTH = 100
MAX_PDF_DIRECT_ARRAY_CHILDREN = 100_000
_PDF_DECODE_PATCH_LOCK = threading.Lock()
REQUIRED_XLSX_ARCHIVE_ENTRIES = frozenset(
    {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
)


class UnsafeQuoteFileError(ValueError):
    """A quote exceeds bounded local parsing resources."""


@dataclass(frozen=True)
class ParsedRow:
    sheet: str | None
    page: int | None
    row: int | None
    cells: str | None
    item_name: str | None
    spec: str | None
    unit: str | None
    quantity: str | None
    unit_price: str | None
    amount: str | None
    maker: str | None
    warnings: tuple[str, ...] = ()


_FIELD_ALIASES = {
    "item_name": (
        "품명",
        "품목",
        "자재명",
        "장치",
        "내용",
        "item",
        "itemname",
        "description",
    ),
    "spec": ("규격", "사양", "spec", "specification", "model"),
    "unit": ("단위", "unit"),
    "quantity": ("수량", "qty", "quantity"),
    "unit_price": ("단가", "unitprice", "price"),
    "amount": ("금액", "합계", "amount", "total"),
    "maker": ("메이커", "제조사", "브랜드", "maker", "manufacturer"),
}
_HEADER_SEPARATORS = re.compile(r"[\s_\-./()\[\]:]+")
_PDF_COLUMNS = re.compile(r"\t+|\s{2,}")


def read_quote(path: Path) -> list[ParsedRow]:
    """Read a supported quote without normalizing its field values."""
    extension = path.suffix.lower()
    if extension == ".xlsx":
        return read_xlsx(path)
    if extension == ".xls":
        return read_xls(path)
    if extension == ".pdf":
        return read_pdf(path)
    raise ValueError(f"unsupported quote extension: {path.suffix}")


def read_xlsx(path: Path) -> list[ParsedRow]:
    _validate_xlsx_archive(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if len(workbook.worksheets) > MAX_XLSX_WORKSHEETS:
            raise UnsafeQuoteFileError("xlsx has too many worksheets")
        parsed: list[ParsedRow] = []
        total_cells = 0
        for sheet in workbook.worksheets:
            if sheet.max_row is None or sheet.max_column is None:
                raise UnsafeQuoteFileError(
                    "xlsx worksheet dimensions are unavailable"
                )
            cells = sheet.max_row * sheet.max_column
            if (
                sheet.max_row > MAX_XLSX_WORKSHEET_ROWS
                or sheet.max_column > MAX_XLSX_WORKSHEET_COLUMNS
                or cells > MAX_XLSX_WORKSHEET_CELLS
            ):
                raise UnsafeQuoteFileError(
                    "xlsx worksheet dimensions exceed safe limits"
                )
            total_cells += cells
            if total_cells > MAX_XLSX_TOTAL_CELLS:
                raise UnsafeQuoteFileError(
                    "xlsx total cell count exceeds safe limits"
                )
            matrix = [
                [cell.value for cell in row]
                for row in sheet.iter_rows()
            ]
            parsed.extend(
                _parse_tabular_rows(
                    matrix,
                    sheet=sheet.title,
                    page=None,
                    row_numbers=True,
                    cell_ranges=True,
                )
            )
        return parsed
    finally:
        workbook.close()


def read_xls(path: Path) -> list[ParsedRow]:
    workbook = xlrd.open_workbook(str(path))
    sheets = workbook.sheets()
    if len(sheets) > MAX_XLS_SHEETS:
        raise UnsafeQuoteFileError("xls has too many worksheets")
    parsed: list[ParsedRow] = []
    total_cells = 0
    for sheet in sheets:
        cells = sheet.nrows * sheet.ncols
        if (
            sheet.nrows > MAX_XLS_SHEET_ROWS
            or sheet.ncols > MAX_XLS_SHEET_COLUMNS
            or cells > MAX_XLS_SHEET_CELLS
        ):
            raise UnsafeQuoteFileError(
                "xls worksheet dimensions exceed safe limits"
            )
        total_cells += cells
        if total_cells > MAX_XLS_TOTAL_CELLS:
            raise UnsafeQuoteFileError(
                "xls total cell count exceeds safe limits"
            )
        matrix = [
            [sheet.cell_value(row, column) for column in range(sheet.ncols)]
            for row in range(sheet.nrows)
        ]
        parsed.extend(
            _parse_tabular_rows(
                matrix,
                sheet=sheet.name,
                page=None,
                row_numbers=True,
                cell_ranges=True,
            )
        )
    return parsed


def _validate_xlsx_archive(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
            raise UnsafeQuoteFileError("xlsx archive has too many entries")
        entry_names = {entry.filename for entry in entries}
        if not REQUIRED_XLSX_ARCHIVE_ENTRIES <= entry_names:
            raise UnsafeQuoteFileError(
                "xlsx archive is missing required workbook entries"
            )
        total_uncompressed = 0
        total_compressed = 0
        for entry in entries:
            total_uncompressed += entry.file_size
            total_compressed += entry.compress_size
            if (
                entry.file_size > 0
                and entry.compress_size == 0
            ):
                raise UnsafeQuoteFileError(
                    "xlsx archive contains an invalid compressed entry"
                )
            if (
                entry.compress_size > 0
                and entry.file_size / entry.compress_size
                > MAX_XLSX_COMPRESSION_RATIO
            ):
                raise UnsafeQuoteFileError(
                    "xlsx archive entry compression ratio is unsafe"
                )
        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise UnsafeQuoteFileError(
                "xlsx archive expands beyond the safe byte limit"
            )
        if (
            total_compressed > 0
            and total_uncompressed / total_compressed
            > MAX_XLSX_COMPRESSION_RATIO
        ):
            raise UnsafeQuoteFileError(
                "xlsx archive compression ratio is unsafe"
            )


def read_pdf(path: Path) -> list[ParsedRow]:
    _preflight_pdf_lexical(path)
    with _bounded_pypdf_flate_decoding():
        return _read_pdf(path)


def _read_pdf(path: Path) -> list[ParsedRow]:
    reader = PdfReader(str(path))
    if len(reader.pages) > MAX_PDF_PAGES:
        raise UnsafeQuoteFileError("pdf has too many pages")
    parsed: list[ParsedRow] = []
    text_total = 0
    row_total = 0
    resource_budget = _PdfResourceBudget()
    for page_number, page in enumerate(reader.pages, start=1):
        _inspect_pdf_page_graph(page, resource_budget)
        text = page.extract_text() or ""
        text_total += len(text)
        if text_total > MAX_PDF_EXTRACTED_TEXT_CHARS:
            raise UnsafeQuoteFileError(
                "pdf extracted text exceeds safe limits"
            )
        lines = text.splitlines()
        row_total += len(lines)
        if row_total > MAX_PDF_EXTRACTED_ROWS:
            raise UnsafeQuoteFileError(
                "pdf extracted row count exceeds safe limits"
            )
        matrix = [
            [part for part in _PDF_COLUMNS.split(line.strip())]
            for line in lines
            if line.strip()
        ]
        parsed.extend(
            _parse_tabular_rows(
                matrix,
                sheet=None,
                page=page_number,
                row_numbers=False,
                cell_ranges=False,
            )
        )
    return parsed


@dataclass
class _PdfResourceBudget:
    visited: set[tuple[object, ...]] = field(default_factory=set)
    object_count: int = 0
    raw_bytes: int = 0
    decoded_bytes: int = 0
    decoded_limit: int = MAX_PDF_DECODED_CONTENT_BYTES
    image_count: int = 0
    image_raw_bytes: int = 0
    image_pixels: int = 0


def _inspect_pdf_page_graph(
    page: Any,
    budget: _PdfResourceBudget,
) -> None:
    if not hasattr(page, "raw_get"):
        return
    for root_name in ("/Contents", "/Resources"):
        try:
            root = page.raw_get(root_name)
        except KeyError:
            continue
        _inspect_pdf_object(root, budget, depth=0)


def _inspect_pdf_object(
    value: Any,
    budget: _PdfResourceBudget,
    *,
    depth: int,
) -> None:
    if depth > MAX_PDF_RESOURCE_DEPTH:
        raise UnsafeQuoteFileError("pdf resource graph is too deep")
    if isinstance(value, IndirectObject):
        marker = (
            "indirect",
            id(value.pdf),
            value.idnum,
            value.generation,
        )
        if marker in budget.visited:
            return
        _visit_pdf_marker(marker, budget)
        _inspect_pdf_object(value.get_object(), budget, depth=depth + 1)
        return
    if isinstance(value, (DictionaryObject, ArrayObject)):
        marker = ("direct", id(value))
        if marker in budget.visited:
            return
        _visit_pdf_marker(marker, budget)
    if isinstance(value, StreamObject):
        _inspect_pdf_stream(value, budget)
        for child in value.values():
            _inspect_pdf_object(child, budget, depth=depth + 1)
    elif isinstance(value, DictionaryObject):
        for child in value.values():
            _inspect_pdf_object(child, budget, depth=depth + 1)
    elif isinstance(value, ArrayObject):
        for child in value:
            _inspect_pdf_object(child, budget, depth=depth + 1)
    else:
        budget.object_count += 1
        if budget.object_count > MAX_PDF_REACHABLE_OBJECTS:
            raise UnsafeQuoteFileError(
                "pdf resource graph has too many primitive children"
            )


def _visit_pdf_marker(
    marker: tuple[object, ...],
    budget: _PdfResourceBudget,
) -> None:
    budget.visited.add(marker)
    budget.object_count += 1
    if budget.object_count > MAX_PDF_REACHABLE_OBJECTS:
        raise UnsafeQuoteFileError(
            "pdf resource graph has too many objects"
        )


def _inspect_pdf_stream(
    stream: StreamObject,
    budget: _PdfResourceBudget,
) -> None:
    raw = stream._data
    budget.raw_bytes += len(raw)
    if budget.raw_bytes > MAX_PDF_COMPRESSED_CONTENT_BYTES:
        raise UnsafeQuoteFileError(
            "pdf reachable stream bytes exceed safe limits"
        )
    filters = _pdf_filter_names(stream)
    if str(stream.get("/Subtype", "")) == "/Image":
        _inspect_pdf_image(stream, filters, raw, budget)
        # PageObject._extract_text explicitly skips image XObjects.
        if filters in (("/DCTDecode",), ("/JPXDecode",)):
            return
    remaining = budget.decoded_limit - budget.decoded_bytes
    if not filters:
        decoded_size = len(raw)
    elif filters in (("/FlateDecode",), ("/Fl",)):
        decoded_size = _bounded_flate_size(raw, remaining)
    else:
        raise UnsafeQuoteFileError(
            "pdf resource uses an unbounded compressed filter"
        )
    budget.decoded_bytes += decoded_size
    if budget.decoded_bytes > budget.decoded_limit:
        raise UnsafeQuoteFileError(
            "pdf decoded resource bytes exceed safe limits"
        )


def _inspect_pdf_image(
    stream: StreamObject,
    filters: tuple[str, ...],
    raw: bytes,
    budget: _PdfResourceBudget,
) -> None:
    width = _pdf_positive_int(stream.get("/Width"))
    height = _pdf_positive_int(stream.get("/Height"))
    if width is None or height is None:
        raise UnsafeQuoteFileError(
            "pdf image dimensions are missing or invalid"
        )
    pixels = width * height
    budget.image_count += 1
    budget.image_raw_bytes += len(raw)
    budget.image_pixels += pixels
    if (
        budget.image_count > MAX_PDF_IMAGE_COUNT
        or budget.image_raw_bytes > MAX_PDF_IMAGE_RAW_BYTES
        or pixels > MAX_PDF_IMAGE_PIXELS
        or budget.image_pixels > MAX_PDF_TOTAL_IMAGE_PIXELS
    ):
        raise UnsafeQuoteFileError("pdf image resources exceed safe limits")
    if filters in (("/DCTDecode",), ("/JPXDecode",)):
        return


def _pdf_positive_int(value: Any) -> int | None:
    if hasattr(value, "get_object"):
        value = value.get_object()
    try:
        parsed = int(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return parsed if parsed > 0 else None


@dataclass(frozen=True)
class _PdfLexicalToken:
    kind: str
    value: str


def _preflight_pdf_lexical(path: Path) -> None:
    data = path.read_bytes()
    if len(data) > MAX_PDF_RAW_BYTES:
        raise UnsafeQuoteFileError("pdf file exceeds safe raw byte limits")
    tokens = _pdf_lexical_tokens(data)
    name_count = sum(token.kind == "name" for token in tokens)
    object_count = sum(
        token.kind == "word" and token.value == "obj"
        for token in tokens
    )
    reference_count = sum(
        token.kind == "word" and token.value == "R"
        for token in tokens
    )
    if name_count > MAX_PDF_LEXICAL_NAMES:
        raise UnsafeQuoteFileError("pdf has too many lexical names")
    if object_count > MAX_PDF_LEXICAL_OBJECTS:
        raise UnsafeQuoteFileError("pdf has too many object declarations")
    if reference_count > MAX_PDF_LEXICAL_REFERENCES:
        raise UnsafeQuoteFileError("pdf has too many indirect references")

    expansion_filters = {
        "RunLengthDecode",
        "RL",
        "LZWDecode",
        "LZW",
    }
    for token in tokens:
        if token.kind == "name" and token.value in expansion_filters:
            raise UnsafeQuoteFileError(
                "pdf declares an unsafe expansion filter"
            )
    for index, token in enumerate(tokens):
        if token.kind == "name" and token.value == "Filter":
            _validate_lexical_filter(tokens, index + 1)


def _validate_lexical_filter(
    tokens: list[_PdfLexicalToken],
    index: int,
) -> None:
    allowed = {"FlateDecode", "Fl", "DCTDecode", "JPXDecode"}
    if index >= len(tokens):
        raise UnsafeQuoteFileError("pdf filter declaration is incomplete")
    value = tokens[index]
    if value.kind == "name":
        if value.value not in allowed:
            raise UnsafeQuoteFileError(
                "pdf declares an unsupported filter"
            )
        return
    if value.kind != "punct" or value.value != "[":
        raise UnsafeQuoteFileError(
            "pdf filter declaration is indirect or malformed"
        )
    names: list[str] = []
    cursor = index + 1
    while cursor < len(tokens):
        current = tokens[cursor]
        if current.kind == "punct" and current.value == "]":
            break
        if current.kind != "name":
            raise UnsafeQuoteFileError(
                "pdf filter array is malformed"
            )
        names.append(current.value)
        cursor += 1
    if cursor >= len(tokens) or len(names) != 1 or names[0] not in allowed:
        raise UnsafeQuoteFileError(
            "pdf filter arrays and chains are unsafe"
        )


def _pdf_lexical_tokens(data: bytes) -> list[_PdfLexicalToken]:
    tokens: list[_PdfLexicalToken] = []
    array_children: list[int] = []
    index = 0
    while index < len(data):
        byte = data[index]
        if byte in b"\x00\t\n\x0c\r ":
            index += 1
            continue
        if byte == ord("%"):
            newline = data.find(b"\n", index + 1)
            index = len(data) if newline < 0 else newline + 1
            continue
        if byte == ord("("):
            index = _skip_pdf_literal_string(data, index + 1)
            continue
        if byte == ord("<") and not data.startswith(b"<<", index):
            closing = data.find(b">", index + 1)
            index = len(data) if closing < 0 else closing + 1
            continue
        if data.startswith(b"<<", index) or data.startswith(b">>", index):
            token = data[index : index + 2].decode("ascii")
            index += 2
            _append_pdf_token(tokens, "punct", token, array_children)
            continue
        if byte in b"[]{}":
            token = chr(byte)
            index += 1
            if token == "[":
                if len(array_children) >= MAX_PDF_LEXICAL_DEPTH:
                    raise UnsafeQuoteFileError(
                        "pdf lexical nesting is too deep"
                    )
                if array_children:
                    array_children[-1] += 1
                array_children.append(0)
            elif token == "]":
                if array_children:
                    array_children.pop()
            _append_pdf_token(tokens, "punct", token, array_children)
            continue
        if byte == ord("/"):
            end = _pdf_token_end(data, index + 1)
            value = _decode_pdf_name(data[index + 1 : end])
            index = end
            _append_pdf_token(tokens, "name", value, array_children)
            continue
        end = _pdf_token_end(data, index)
        if end == index:
            index += 1
            continue
        value = data[index:end].decode("latin-1")
        index = end
        _append_pdf_token(tokens, "word", value, array_children)
        if value == "stream":
            closing = data.find(b"endstream", index)
            index = len(data) if closing < 0 else closing + len(b"endstream")
    return tokens


def _append_pdf_token(
    tokens: list[_PdfLexicalToken],
    kind: str,
    value: str,
    array_children: list[int],
) -> None:
    tokens.append(_PdfLexicalToken(kind, value))
    if len(tokens) > MAX_PDF_LEXICAL_TOKENS:
        raise UnsafeQuoteFileError("pdf has too many lexical tokens")
    if array_children and not (kind == "punct" and value in {"[", "]"}):
        array_children[-1] += 1
        if array_children[-1] > MAX_PDF_DIRECT_ARRAY_CHILDREN:
            raise UnsafeQuoteFileError(
                "pdf direct array has too many children"
            )


def _skip_pdf_literal_string(data: bytes, index: int) -> int:
    depth = 1
    while index < len(data) and depth:
        byte = data[index]
        if byte == ord("\\"):
            index += 2
            continue
        if byte == ord("("):
            depth += 1
            if depth > MAX_PDF_LEXICAL_DEPTH:
                raise UnsafeQuoteFileError(
                    "pdf literal string nesting is too deep"
                )
        elif byte == ord(")"):
            depth -= 1
        index += 1
    return index


def _pdf_token_end(data: bytes, index: int) -> int:
    delimiters = b"\x00\t\n\x0c\r ()<>[]{}/%"
    while index < len(data) and data[index] not in delimiters:
        index += 1
    return index


def _decode_pdf_name(value: bytes) -> str:
    decoded = bytearray()
    index = 0
    while index < len(value):
        if (
            value[index] == ord("#")
            and index + 2 < len(value)
            and all(
                character in b"0123456789abcdefABCDEF"
                for character in value[index + 1 : index + 3]
            )
        ):
            decoded.append(int(value[index + 1 : index + 3], 16))
            index += 3
        else:
            decoded.append(value[index])
            index += 1
    return decoded.decode("latin-1")


def _bounded_pdf_page_content(
    page: Any,
    *,
    decoded_remaining: int,
) -> tuple[int, int]:
    budget = _PdfResourceBudget(decoded_limit=decoded_remaining)
    if not hasattr(page, "raw_get"):
        return 0, 0
    try:
        contents = page.raw_get("/Contents")
    except KeyError:
        return 0, 0
    _inspect_pdf_object(contents, budget, depth=0)
    return budget.raw_bytes, budget.decoded_bytes


def _pdf_filter_names(stream: StreamObject) -> tuple[str, ...]:
    value = stream.get("/Filter")
    if value is None:
        return ()
    value = value.get_object() if hasattr(value, "get_object") else value
    values = value if isinstance(value, ArrayObject) else [value]
    return tuple(str(item.get_object()) for item in values)


def _bounded_flate_size(data: bytes, maximum: int) -> int:
    return len(_bounded_flate_decode(data, maximum))


def _bounded_flate_decode(data: bytes, maximum: int) -> bytes:
    for window_bits in (zlib.MAX_WBITS, zlib.MAX_WBITS | 32):
        decoder = zlib.decompressobj(window_bits)
        try:
            decoded = decoder.decompress(data, maximum + 1)
            if len(decoded) > maximum or decoder.unconsumed_tail:
                raise UnsafeQuoteFileError(
                    "pdf decoded content exceeds safe limits"
                )
            decoded += decoder.flush(maximum + 1 - len(decoded))
            if len(decoded) > maximum:
                raise UnsafeQuoteFileError(
                    "pdf decoded content exceeds safe limits"
                )
            return decoded
        except zlib.error:
            continue
    raise UnsafeQuoteFileError("pdf flate content cannot be decoded safely")


@contextmanager
def _bounded_pypdf_flate_decoding():
    """Serialize pypdf's process-global Flate hook and restore it reliably."""

    import pypdf.filters as pdf_filters

    with _PDF_DECODE_PATCH_LOCK:
        original = pdf_filters.decompress
        remaining = MAX_PDF_TOTAL_FLATE_DECODED_BYTES

        def bounded_decompress(data: bytes) -> bytes:
            nonlocal remaining
            decoded = _bounded_flate_decode(data, remaining)
            remaining -= len(decoded)
            return decoded

        pdf_filters.decompress = bounded_decompress
        try:
            yield
        finally:
            pdf_filters.decompress = original


def _parse_tabular_rows(
    rows: list[list[Any]],
    *,
    sheet: str | None,
    page: int | None,
    row_numbers: bool,
    cell_ranges: bool,
) -> list[ParsedRow]:
    header_index, columns = _find_header(rows)
    if header_index is None:
        return (
            _parse_fixed_column_fallback(rows, sheet=sheet, page=page)
            if cell_ranges
            else []
        )

    parsed: list[ParsedRow] = []
    mapped_columns = sorted(columns.values())
    for row_index, values in enumerate(
        rows[header_index + 1 :],
        start=header_index + 1,
    ):
        fields = {
            field: _raw_text(
                values[column] if column < len(values) else None
            )
            for field, column in columns.items()
        }
        if not any(fields.values()):
            continue

        source_row = row_index + 1 if row_numbers else None
        source_cells = None
        if cell_ranges:
            first_column = get_column_letter(mapped_columns[0] + 1)
            last_column = get_column_letter(mapped_columns[-1] + 1)
            source_cells = (
                f"{first_column}{source_row}:{last_column}{source_row}"
            )
        parsed.append(
            ParsedRow(
                sheet=sheet,
                page=page,
                row=source_row,
                cells=source_cells,
                item_name=fields.get("item_name"),
                spec=fields.get("spec"),
                unit=fields.get("unit"),
                quantity=fields.get("quantity"),
                unit_price=fields.get("unit_price"),
                amount=fields.get("amount"),
                maker=fields.get("maker"),
            )
        )
    return parsed


def _find_header(
    rows: list[list[Any]],
) -> tuple[int | None, dict[str, int]]:
    for row_index, row in enumerate(rows):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(row):
            field = _field_for_header(value)
            if field is not None and field not in columns:
                columns[field] = column_index
        if "item_name" not in columns and (
            "unit_price" in columns or "amount" in columns
        ):
            item_column = _fallback_item_column(row, columns)
            if item_column is not None:
                columns["item_name"] = item_column
        has_item = "item_name" in columns
        has_price = (
            "unit_price" in columns or "amount" in columns
        )
        if has_item and has_price:
            return row_index, columns
    return None, {}


def _field_for_header(value: Any) -> str | None:
    normalized = _HEADER_SEPARATORS.sub(
        "",
        _raw_text(value) or "",
    ).casefold()
    if not normalized:
        return None
    for field, aliases in _FIELD_ALIASES.items():
        if normalized in aliases:
            return field
    undecorated = normalized.removesuffix("원").removesuffix("krw")
    for field in ("unit_price", "amount"):
        if undecorated in _FIELD_ALIASES[field]:
            return field
    return None


def _fallback_item_column(
    row: list[Any],
    columns: dict[str, int],
) -> int | None:
    price_column = columns.get("unit_price", columns.get("amount"))
    if price_column is None:
        return None
    mapped_columns = set(columns.values())
    ignored = {"단위", "수량", "구분", "번호", "순번", "위치"}
    for column in range(price_column - 1, -1, -1):
        value = _HEADER_SEPARATORS.sub(
            "",
            _raw_text(row[column]) or "",
        ).casefold()
        if (
            value
            and column not in mapped_columns
            and not any(word in value for word in ignored)
        ):
            return column
    return None


def _raw_text(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_fixed_column_fallback(
    rows: list[list[Any]],
    *,
    sheet: str | None,
    page: int | None,
) -> list[ParsedRow]:
    """Conservatively recognize the observed C/E/F/H headerless layout."""
    parsed: list[ParsedRow] = []
    for row_index, values in enumerate(rows, start=1):
        if len(values) < 8:
            continue
        item_name = _raw_text(values[2])
        quantity = _raw_text(values[4])
        unit = _raw_text(values[5])
        unit_price = _raw_text(values[7])
        if not _is_safe_fixed_column_row(
            item_name,
            quantity,
            unit,
            unit_price,
        ):
            continue
        parsed.append(
            ParsedRow(
                sheet=sheet,
                page=page,
                row=row_index,
                cells=f"C{row_index}:H{row_index}",
                item_name=item_name,
                spec=None,
                unit=unit,
                quantity=quantity,
                unit_price=unit_price,
                amount=None,
                maker=None,
                warnings=("FALLBACK_FIXED_C_E_F_H",),
            )
        )
    return parsed


def _is_safe_fixed_column_row(
    item_name: str | None,
    quantity: str | None,
    unit: str | None,
    unit_price: str | None,
) -> bool:
    item = (item_name or "").strip()
    normalized_unit = (unit or "").strip()
    return bool(
        len(item) >= 2
        and not _is_number(item)
        and normalized_unit
        and len(normalized_unit) <= 20
        and _positive_number(quantity)
        and _positive_number(unit_price)
    )


def _positive_number(value: str | None) -> bool:
    try:
        return Decimal((value or "").replace(",", "").strip()) > 0
    except InvalidOperation:
        return False


def _is_number(value: str) -> bool:
    try:
        Decimal(value.replace(",", "").strip())
        return True
    except InvalidOperation:
        return False
