"""
협력사 견적 단가 AI 비교·분석 시스템 — MVP
현대위아 구매본부
"""
import streamlit as st
import pandas as pd
import numpy as np
import json
import re
import io
from pathlib import Path
import openpyxl
import xlrd

# ── 페이지 설정 ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="견적 단가 AI 분석 시스템 | 현대위아 구매본부",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 상수 ─────────────────────────────────────────────────────────────────────
DB_PATH = Path(__file__).parent / "표준단가DB.xlsx"
THRESHOLD = 0.15  # 이상가 임계값 15%

# ── 캐시: 표준단가 DB 로드 (파일 수정시간 기반 자동 무효화) ─────────────────
@st.cache_data
def load_standard_db(mtime: float = 0.0):
    df = pd.read_excel(DB_PATH, sheet_name="표준단가DB", header=2)
    df.columns = df.columns.str.strip()
    # 숫자 컬럼 정제
    for col in ['단가_최저(원)', '단가_평균(원)', '단가_최고(원)', '데이터건수']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df = df[df['품명'].notna() & (df['품명'] != '') & (df['단가_평균(원)'] > 0)]
    df = df.reset_index(drop=True)
    return df

# ── AI 유사도 검색 (TF-IDF 기반, 경량 모델) ──────────────────────────────
@st.cache_resource
def build_search_index(df):
    from sklearn.feature_extraction.text import TfidfVectorizer
    # 품명+규격 합성 텍스트
    df['검색텍스트'] = (df['품명'].fillna('') + ' ' + df['규격'].fillna('')).str.strip()
    vectorizer = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 4), min_df=1)
    matrix = vectorizer.fit_transform(df['검색텍스트'])
    return vectorizer, matrix

def find_similar(query: str, df: pd.DataFrame, vectorizer, matrix, top_n: int = 5):
    """품명 쿼리에 대해 유사 품목 Top-N 반환"""
    from sklearn.metrics.pairwise import cosine_similarity
    if not query or not query.strip():
        return pd.DataFrame()
    try:
        q_vec = vectorizer.transform([query.strip()])
        sims = cosine_similarity(q_vec, matrix).flatten()
        top_idx = np.argsort(sims)[::-1][:top_n]
        result = df.iloc[top_idx].copy()
        result['유사도'] = sims[top_idx]
        result = result[result['유사도'] > 0.05]  # 최소 유사도 필터
        return result.reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

# ── 견적서 파싱 ───────────────────────────────────────────────────────────
def parse_quotation_excel(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """견적서 Excel에서 품목·단가 추출"""
    items = []

    try:
        if filename.endswith('.xls'):
            wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override='cp949')
            sheets = wb.sheet_names()
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets = wb.sheetnames
    except Exception as e:
        st.error(f"파일 읽기 오류: {e}")
        return pd.DataFrame()

    for sh_name in sheets:
        try:
            if filename.endswith('.xls'):
                ws = wb.sheet_by_name(sh_name)
                rows = []
                for r in range(ws.nrows):
                    rows.append([str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)])
            else:
                ws = wb[sh_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(v).strip() if v is not None else '' for v in row])
        except Exception:
            continue

        # 단위장비 / 갑지 시트 우선, 단위공사별도 처리
        is_item_sheet = any(k in sh_name for k in [
            '단위장비', 'HVAC조립', 'ASSY', 'ASS', '성형기', '압입기', '도포기',
            '포밍기', '조립기', '검사기', '이재', '컨베어', '헤어핀', '코어',
            '절연', '인서팅', '확관', '트위스팅', '하부', '서브', 'OTHER',
            '전기', 'VISION', '포승', 'U핀', 'I핀',
        ])
        is_summary = '갑지' in sh_name or '원가항목' in sh_name or '단위공사별' in sh_name

        if not (is_item_sheet or is_summary):
            continue

        unit_name = sh_name
        for row in rows[:3]:
            clean = [c for c in row if c.strip() and '단위공사명' not in c and '(첨부)' not in c]
            if clean and len(clean[0]) > 2 and not clean[0].startswith('□'):
                unit_name = clean[0]
                break

        for row in rows:
            clean = [c for c in row if c.strip() and c != 'None']
            nums = []
            texts = []
            for c in clean:
                try:
                    v = float(c.replace(',', ''))
                    nums.append(v)
                except Exception:
                    texts.append(c)

            if len(nums) >= 2 and len(texts) >= 1:
                try:
                    금액 = nums[-1]
                    단가 = nums[-2]
                    수량 = nums[-3] if len(nums) >= 3 else None
                    if 단가 > 0 and 금액 >= 1000:
                        품명 = texts[0] if texts else ""
                        규격 = texts[1] if len(texts) > 1 else ""
                        단위 = texts[2] if len(texts) > 2 else ""
                        메이커 = texts[-1] if len(texts) > 3 else ""
                        skip_words = ['소계', '합계', '이윤', '관리비', '노무비', '경비', 'M/D',
                                      '철거', '양산', '설계', '제작', '조립', '설치', '시운전', '도장',
                                      '포장', '운송', '안전', '이윤', '순번', '구분', '원가', '합 계']
                        if any(k in 품명 for k in skip_words):
                            continue
                        if 품명 in ['', '0'] or 품명.startswith('0'):
                            continue
                        items.append({
                            '단위장비명': unit_name,
                            '품명': 품명,
                            '규격': 규격,
                            '단위': 단위,
                            '수량': 수량,
                            '견적단가(원)': int(단가),
                            '견적금액(원)': int(금액),
                            '메이커': 메이커,
                        })
                except Exception:
                    pass

    if not items:
        return pd.DataFrame()

    df = pd.DataFrame(items).drop_duplicates(subset=['품명', '규격', '견적단가(원)'])
    return df.reset_index(drop=True)

# ── 비교 분석 ─────────────────────────────────────────────────────────────
def analyze_items(quote_df: pd.DataFrame, db_df: pd.DataFrame, vectorizer, matrix) -> pd.DataFrame:
    results = []
    progress = st.progress(0)
    total = len(quote_df)

    for idx, row in quote_df.iterrows():
        progress.progress((idx + 1) / total)
        query = f"{row['품명']} {row.get('규격', '')}".strip()
        sim_df = find_similar(query, db_df, vectorizer, matrix, top_n=1)

        std_avg = None
        std_min = None
        std_max = None
        std_item = None
        similarity = None
        판정 = "표준단가 없음"
        차이율 = None
        이상여부 = False

        if not sim_df.empty and sim_df.iloc[0]['유사도'] > 0.1:
            best = sim_df.iloc[0]
            std_avg = int(best['단가_평균(원)'])
            std_min = int(best['단가_최저(원)'])
            std_max = int(best['단가_최고(원)'])
            std_item = best['품명']
            similarity = round(float(best['유사도']) * 100, 1)
            견적가 = row['견적단가(원)']
            차이율 = round((견적가 - std_avg) / std_avg * 100, 1) if std_avg > 0 else None
            if 차이율 is not None:
                if 차이율 > THRESHOLD * 100:
                    판정 = f"⚠️ 이상가 (+{차이율:.1f}%)"
                    이상여부 = True
                elif 차이율 < -THRESHOLD * 100:
                    판정 = f"✅ 적정 (저렴 {차이율:.1f}%)"
                else:
                    판정 = f"✅ 적정 ({차이율:+.1f}%)"

        results.append({
            '단위장비명': row.get('단위장비명', ''),
            '품명': row['품명'],
            '규격': row.get('규격', ''),
            '단위': row.get('단위', ''),
            '수량': row.get('수량'),
            '견적단가(원)': row['견적단가(원)'],
            '표준단가_평균(원)': std_avg,
            '표준단가_최저(원)': std_min,
            '표준단가_최고(원)': std_max,
            '매핑품명': std_item,
            '유사도(%)': similarity,
            '차이율(%)': 차이율,
            '판정': 판정,
            '이상여부': 이상여부,
        })

    progress.empty()
    return pd.DataFrame(results)

# ── KPI 검색 (subprocess 격리 — Playwright 스레드 충돌 방지) ──────────────
def kpi_search_subprocess(keyword: str, max_results: int = 5) -> list:
    """kpi_worker.py를 별도 프로세스로 실행해 스레드 격리"""
    import subprocess, json as _json
    worker = Path(__file__).parent / "kpi_worker.py"
    try:
        r = subprocess.run(
            ["python", str(worker), keyword, str(max_results)],
            capture_output=True, timeout=45,
            cwd=str(Path(__file__).parent)
        )
        text = r.stdout.decode("utf-8", errors="replace").strip()
        if not text:
            return []
        data = _json.loads(text)
        if isinstance(data, dict) and "error" in data:
            return []
        return data if isinstance(data, list) else []
    except Exception:
        return []


def kpi_login_check() -> bool:
    """KPI 로그인 가능 여부를 subprocess로 확인"""
    import subprocess, json as _json
    worker = Path(__file__).parent / "kpi_worker.py"
    try:
        r = subprocess.run(
            ["python", str(worker), "__login_test__", "1"],
            capture_output=True, timeout=30,
            cwd=str(Path(__file__).parent)
        )
        text = r.stdout.decode("utf-8", errors="replace").strip()
        data = _json.loads(text) if text else {}
        return not (isinstance(data, dict) and "error" in data)
    except Exception:
        return False


# ── 검토 로직 함수 ────────────────────────────────────────────────────────
def compute_final_price(row) -> int:
    """검토근거 선택에 따라 최종검토가 반환"""
    basis = row.get("검토근거", "")
    if basis == "견적가 적용":
        return int(row["견적단가"])
    elif basis == "표준단가 적용":
        v = row.get("표준단가", 0)
        return int(v) if v else int(row["견적단가"])
    elif basis == "시장가 적용":
        v = row.get("KPI시장가", 0)
        return int(v) if v else int(row["할인율적용가"])
    else:  # 할인율 적용
        return int(row["할인율적용가"])


def build_review_df(result_df: pd.DataFrame, kpi_dict: dict, default_discount: float) -> pd.DataFrame:
    """result_df + KPI 결과로 검토 DataFrame 생성"""
    rows = []
    for _, r in result_df.iterrows():
        has_db = (pd.notna(r.get("표준단가_평균(원)")) and
                  (r.get("유사도(%)") or 0) >= 20)
        견적가 = int(r["견적단가(원)"])
        표준가 = int(r["표준단가_평균(원)"]) if has_db else 0
        할인율 = default_discount
        할인가 = int(견적가 * (1 - 할인율 / 100))
        kpi가 = int(kpi_dict.get(r["품명"], 0))

        if 표준가 > 0:
            방향 = "↓ 견적보다 낮음" if 표준가 < 견적가 else "↑ 견적보다 높음"
        else:
            방향 = "N/A"

        # 시스템 추천 로직
        if has_db and 표준가 > 0:
            if 견적가 <= 표준가:
                추천 = "견적가 적용"
            else:
                추천 = "표준단가 적용"
        else:
            if kpi가 > 0 and kpi가 < 할인가:
                추천 = "시장가 적용"
            else:
                추천 = "할인율 적용"

        row_dict = {
            "단위장비명": r.get("단위장비명", ""),
            "품명": r["품명"],
            "규격": r.get("규격", ""),
            "단위": r.get("단위", ""),
            "수량": r.get("수량") or 1,
            "견적단가": 견적가,
            "DB존재": "✅" if has_db else "❌",
            "표준단가": 표준가,
            "표준단가_방향": 방향,
            "KPI시장가": kpi가,
            "할인율(%)": 할인율,
            "할인율적용가": 할인가,
            "시스템추천": 추천,
            "검토근거": 추천,   # 편집 가능
        }
        row_dict["최종검토가"] = compute_final_price(row_dict)
        rows.append(row_dict)
    return pd.DataFrame(rows)


def save_review_to_db(reviewed_df: pd.DataFrame, source_file: str, base_dir: str) -> tuple:
    """
    검토 결과를 reviewed_db.json에 저장하고,
    DB 미매칭 품목을 표준단가DB_집계.json에 추가 후 Excel 재생성.
    반환: (성공여부, 메시지)
    """
    import json, subprocess
    from datetime import date
    base = Path(base_dir)
    today = str(date.today())

    # 1. reviewed_db.json append
    reviewed_path = base / "reviewed_db.json"
    existing = []
    if reviewed_path.exists():
        try:
            existing = json.loads(reviewed_path.read_text(encoding="utf-8"))
        except Exception:
            existing = []

    new_records = []
    for _, r in reviewed_df.iterrows():
        new_records.append({
            "검토일": today,
            "출처파일": source_file,
            "단위장비명": r.get("단위장비명", ""),
            "품명": r["품명"],
            "규격": r.get("규격", ""),
            "단위": r.get("단위", ""),
            "수량": float(r.get("수량") or 1),
            "견적단가": int(r["견적단가"]),
            "최종검토가": int(r["최종검토가"]),
            "검토근거": r["검토근거"],
            "표준단가": int(r.get("표준단가") or 0),
            "KPI시장가": int(r.get("KPI시장가") or 0),
            "할인율": float(r.get("할인율(%)") or 0),
        })
    reviewed_path.write_text(
        json.dumps(existing + new_records, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # 2. DB 미매칭 품목 → 표준단가DB_집계.json에 신규 추가
    agg_path = base / "표준단가DB_집계.json"
    try:
        agg = json.loads(agg_path.read_text(encoding="utf-8"))
    except Exception:
        agg = []

    existing_keys = {f"{r['품명']}||{r['규격']}" for r in agg}
    next_id = len(agg) + 1
    added = 0
    for r in new_records:
        if r.get("표준단가", 0) == 0:
            key = f"{r['품명']}||{r['규격']}"
            if key not in existing_keys and r["최종검토가"] > 0:
                agg.append({
                    "표준품목ID": f"STD-{next_id:04d}",
                    "품목분류": "설비",
                    "품명": r["품명"],
                    "규격": r["규격"],
                    "단위": r["단위"],
                    "단가_최저": r["최종검토가"],
                    "단가_평균": r["최종검토가"],
                    "단가_최고": r["최종검토가"],
                    "데이터건수": 1,
                    "주요메이커": "",
                    "최근견적일": today,
                    "출처": f"검토확정({source_file})",
                })
                existing_keys.add(key)
                next_id += 1
                added += 1

    agg_path.write_text(json.dumps(agg, ensure_ascii=False, indent=2), encoding="utf-8")

    # 3. Excel 재생성
    try:
        subprocess.run(
            ["python", str(base / "build_db.py")],
            capture_output=True, timeout=60
        )
    except Exception:
        pass

    return True, f"저장 완료 — {len(new_records)}건 검토 기록, {added}개 신규 품목 DB 추가"


# ── UI: 스타일 ────────────────────────────────────────────────────────────
st.markdown("""
<style>
.main-title { font-size: 1.8rem; font-weight: 700; color: #1F3864; }
.sub-title  { font-size: 1rem; color: #666; margin-bottom: 1rem; }
.metric-box { background: #EBF3FB; border-radius: 8px; padding: 12px 16px; text-align: center; }
.metric-val { font-size: 1.6rem; font-weight: 700; color: #1F3864; }
.metric-lbl { font-size: 0.8rem; color: #666; }
.alert-box  { background: #FFF3CD; border-left: 4px solid #FFA500; padding: 10px 14px; border-radius: 4px; }
.ok-box     { background: #D4EDDA; border-left: 4px solid #28A745; padding: 10px 14px; border-radius: 4px; }
</style>
""", unsafe_allow_html=True)

# ── 사이드바 ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/4/4e/Hyundai_Motor_Group_CI.svg/200px-Hyundai_Motor_Group_CI.svg.png", width=120)
    st.markdown("### 🔧 설정")
    threshold_pct = st.slider("이상가 임계값 (%)", 5, 30, 15, step=5,
                               help="표준단가 대비 견적가 초과 시 이상가로 판정")
    top_n = st.slider("유사 품목 표시 수", 1, 10, 5)
    discount_rate_default = st.number_input(
        "기본 할인율 (%)", 0.0, 50.0, 23.8, step=0.1,
        help="표준단가DB 미매칭 품목에 적용할 기본 할인율 (검토표에서 품목별로 수정 가능)"
    )
    st.markdown("---")
    st.markdown("**표준단가 DB 현황**")
    if st.button("🔄 캐시 초기화"):
        st.cache_data.clear()
        st.cache_resource.clear()
        st.rerun()
    try:
        _mtime = DB_PATH.stat().st_mtime if DB_PATH.exists() else 0.0
        db_df = load_standard_db(mtime=_mtime)
        st.success(f"✅ {len(db_df):,}개 품목 로드됨")
    except Exception as e:
        st.error(f"DB 로드 실패: {e}")
        db_df = pd.DataFrame()
    st.markdown("---")
    st.markdown("**🌐 KPI 시장가 연동**")
    kpi_enabled = st.toggle("한국물가정보 조회", value=False,
                             help="kpi.or.kr 로그인 후 시장가를 자동으로 가져옵니다")
    if kpi_enabled:
        if "kpi_logged_in" not in st.session_state:
            st.session_state.kpi_logged_in = False
        if not st.session_state.kpi_logged_in:
            with st.spinner("KPI 연결 확인 중..."):
                ok = kpi_login_check()
                st.session_state.kpi_logged_in = ok
                if ok:
                    st.success("✅ KPI 연결 확인")
                else:
                    st.error("❌ KPI 접속 실패 (네트워크 또는 ID/PW 확인)")
                    kpi_enabled = False
        else:
            st.success("✅ KPI 연결됨")
    st.markdown("---")
    st.caption("현대위아 구매본부 | MVP v1.0")

THRESHOLD = threshold_pct / 100

# ── 메인 ─────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">🔍 협력사 견적 단가 AI 비교·분석 시스템</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">견적서를 업로드하면 AI가 표준단가와 자동 비교하고 이상가를 탐지합니다</div>', unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["📤 견적서 분석", "🗄️ 표준단가 DB 조회", "🔎 품목 단가 검색"])

# ── Tab 1: 견적서 분석 ──────────────────────────────────────────────────
with tab1:
    col_up, col_info = st.columns([2, 1])
    with col_up:
        uploaded = st.file_uploader(
            "견적서 파일 업로드 (Excel: .xlsx / .xls)",
            type=['xlsx', 'xls'],
            help="현대/기아차 설비협력업체 견적통일양식 또는 유사 형식 지원"
        )

    if uploaded:
        with col_info:
            st.info(f"**파일명**: {uploaded.name}\n\n**크기**: {uploaded.size:,} bytes")

        with st.spinner("견적서 파싱 중..."):
            file_bytes = uploaded.read()
            quote_df = parse_quotation_excel(file_bytes, uploaded.name)

        if quote_df.empty:
            st.error("❌ 품목 데이터를 추출할 수 없습니다. 파일 형식을 확인해 주세요.")
        else:
            st.success(f"✅ **{len(quote_df)}개 품목** 추출 완료")

            with st.expander("📋 추출된 품목 목록 (원본)", expanded=False):
                st.dataframe(
                    quote_df.style.format({'견적단가(원)': '{:,.0f}', '견적금액(원)': '{:,.0f}'}),
                    use_container_width=True, height=300
                )

            if db_df.empty:
                st.warning("표준단가 DB가 로드되지 않아 비교를 수행할 수 없습니다.")
            else:
                st.markdown("---")
                st.markdown("### 🤖 AI 유사도 분석 & 표준단가 비교")

                with st.spinner("AI 유사도 분석 중... (처음 실행 시 모델 초기화 20초 소요)"):
                    vectorizer, matrix = build_search_index(db_df)
                    result_df = analyze_items(quote_df, db_df, vectorizer, matrix)

                # KPI 박스
                total = len(result_df)
                matched = result_df['표준단가_평균(원)'].notna().sum()
                abnormal = result_df['이상여부'].sum()
                total_quote = result_df['견적단가(원)'].sum()
                total_std = result_df['표준단가_평균(원)'].fillna(result_df['견적단가(원)']).sum()
                saving = total_std * (THRESHOLD if total_std > 0 else 0)

                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.markdown(f'<div class="metric-box"><div class="metric-val">{total}</div><div class="metric-lbl">전체 품목 수</div></div>', unsafe_allow_html=True)
                with c2:
                    st.markdown(f'<div class="metric-box"><div class="metric-val">{matched}</div><div class="metric-lbl">표준단가 매핑</div></div>', unsafe_allow_html=True)
                with c3:
                    color = "#C0392B" if abnormal > 0 else "#27AE60"
                    st.markdown(f'<div class="metric-box"><div class="metric-val" style="color:{color}">{abnormal}</div><div class="metric-lbl">⚠️ 이상가 품목</div></div>', unsafe_allow_html=True)
                with c4:
                    st.markdown(f'<div class="metric-box"><div class="metric-val">{total_quote:,.0f}원</div><div class="metric-lbl">견적 합계</div></div>', unsafe_allow_html=True)

                st.markdown("---")

                # 비교 결과 테이블
                st.markdown("#### 📊 품목별 비교 결과")

                display_df = result_df[[
                    '품명', '규격', '단위', '수량',
                    '견적단가(원)', '표준단가_평균(원)', '차이율(%)', '판정', '매핑품명', '유사도(%)'
                ]].copy()

                def highlight_rows(row):
                    if row.get('판정', '').startswith('⚠️'):
                        return ['background-color: #FFDEDE'] * len(row)
                    elif row.get('판정', '').startswith('✅'):
                        return ['background-color: #D4F5DC'] * len(row)
                    return [''] * len(row)

                styled = display_df.style\
                    .apply(highlight_rows, axis=1)\
                    .format({
                        '견적단가(원)': lambda x: f'{int(x):,}' if pd.notna(x) else '-',
                        '표준단가_평균(원)': lambda x: f'{int(x):,}' if pd.notna(x) else '-',
                        '차이율(%)': lambda x: f'{x:+.1f}%' if pd.notna(x) else '-',
                        '유사도(%)': lambda x: f'{x:.1f}%' if pd.notna(x) else '-',
                        '수량': lambda x: f'{x:.0f}' if pd.notna(x) else '-',
                    })

                st.dataframe(styled, use_container_width=True, height=500)

                # 이상가 품목 별도 강조
                abnormal_df = result_df[result_df['이상여부']]
                if not abnormal_df.empty:
                    st.markdown("---")
                    st.markdown(f"#### ⚠️ 이상가 품목 목록 ({len(abnormal_df)}건 — 표준단가 대비 +{threshold_pct}% 초과)")
                    for _, r in abnormal_df.iterrows():
                        견적 = r['견적단가(원)']
                        표준 = r['표준단가_평균(원)']
                        차이 = r['차이율(%)']
                        협상금액 = int((견적 - 표준) * (r['수량'] if pd.notna(r['수량']) else 1)) if pd.notna(표준) else 0
                        st.markdown(f"""
<div class="alert-box">
<b>{r['품명']}</b> ({r['규격']})<br>
견적단가 <b>{견적:,}원</b> → 표준단가 <b>{int(표준):,}원</b> | 차이율 <b style="color:red">{차이:+.1f}%</b> | 협상 가능 금액 약 <b>{협상금액:,}원</b>
</div>
""", unsafe_allow_html=True)

                # Excel 다운로드
                st.markdown("---")
                st.markdown("#### 📥 리포트 다운로드")

                def make_excel_report(result_df, uploaded_name):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "비교분석결과"
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    hfont = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
                    hfill = PatternFill('solid', fgColor='1F3864')
                    warn_fill = PatternFill('solid', fgColor='FFDEDE')
                    ok_fill = PatternFill('solid', fgColor='D4F5DC')
                    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    right = Alignment(horizontal='right', vertical='center')

                    headers = ['단위장비명', '품명', '규격', '단위', '수량',
                               '견적단가(원)', '표준단가_최저(원)', '표준단가_평균(원)', '표준단가_최고(원)',
                               '차이율(%)', '판정', '매핑품명', '유사도(%)']
                    widths = [20, 30, 25, 8, 8, 15, 15, 15, 15, 12, 20, 25, 12]

                    ws.cell(1, 1, f"협력사 견적 단가 비교분석 리포트 — {uploaded_name}").font = Font(name='맑은 고딕', bold=True, size=13, color='1F3864')
                    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
                    ws.row_dimensions[1].height = 24

                    for ci, (h, w) in enumerate(zip(headers, widths), 1):
                        cell = ws.cell(2, ci, h)
                        cell.font = hfont
                        cell.fill = hfill
                        cell.alignment = center
                        ws.column_dimensions[get_column_letter(ci)].width = w
                    ws.row_dimensions[2].height = 22

                    for ri, row in result_df.iterrows():
                        vals = [
                            row.get('단위장비명', ''), row['품명'], row.get('규격', ''),
                            row.get('단위', ''), row.get('수량'),
                            row['견적단가(원)'], row.get('표준단가_최저(원)'),
                            row.get('표준단가_평균(원)'), row.get('표준단가_최고(원)'),
                            row.get('차이율(%)'), row.get('판정', '표준단가 없음'),
                            row.get('매핑품명', ''), row.get('유사도(%)')
                        ]
                        is_warn = row.get('이상여부', False)
                        fill = warn_fill if is_warn else None

                        for ci, val in enumerate(vals, 1):
                            cell = ws.cell(ri + 3, ci, val)
                            cell.font = Font(name='맑은 고딕', size=9)
                            cell.alignment = center if ci in [1, 4, 5, 10, 11, 13] else right if ci in [6, 7, 8, 9] else Alignment(horizontal='left', vertical='center')
                            if fill:
                                cell.fill = fill

                        # 숫자 포맷
                        for ci in [6, 7, 8, 9]:
                            c = ws.cell(ri + 3, ci)
                            c.number_format = '#,##0'
                        ws.cell(ri + 3, 10).number_format = '+0.0%;-0.0%'
                        ws.row_dimensions[ri + 3].height = 18

                    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(result_df)}"
                    ws.freeze_panes = 'A3'

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf

                report_buf = make_excel_report(result_df, uploaded.name)
                st.download_button(
                    label="📥 Excel 비교분석 리포트 다운로드",
                    data=report_buf,
                    file_name=f"견적비교분석_{uploaded.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # ─────────────────────────────────────────────────────────
                # 견적 검토 및 최적가 결정
                # ─────────────────────────────────────────────────────────
                st.markdown("---")
                st.markdown("## 📋 견적 검토 및 최적가 결정")
                st.caption("시스템 추천을 확인하고 '검토근거'를 변경하거나 '할인율(%)'을 수정하세요. 완료 후 [검토 확정 & DB 저장]을 클릭하세요.")

                # KPI 시장가 일괄 조회 버튼
                kpi_dict = st.session_state.get("kpi_prices", {})
                no_db_items = result_df[
                    result_df["표준단가_평균(원)"].isna() |
                    (result_df["유사도(%)"].fillna(0) < 20)
                ]["품명"].tolist()

                col_kpi_btn, col_kpi_status = st.columns([1, 3])
                with col_kpi_btn:
                    kpi_avail = kpi_enabled and st.session_state.get("kpi_logged_in", False)
                    if st.button(
                        "🌐 KPI 시장가 일괄 조회",
                        disabled=not kpi_avail or not no_db_items,
                        help="DB 미매칭 품목의 KPI 시장가를 조회합니다 (KPI 연동 ON 필요)"
                    ):
                        kpi_result = {}
                        prog = st.progress(0)
                        status_txt = st.empty()
                        for i, nm in enumerate(no_db_items):
                            status_txt.caption(f"조회 중: {nm} ({i+1}/{len(no_db_items)})")
                            items = kpi_search_subprocess(nm.split()[0], max_results=3)
                            if items:
                                kpi_result[nm] = items[0]["단가"]
                            prog.progress((i + 1) / len(no_db_items))
                        prog.empty()
                        status_txt.empty()
                        st.session_state["kpi_prices"] = kpi_result
                        st.rerun()
                with col_kpi_status:
                    if kpi_dict:
                        matched_kpi = sum(1 for n in no_db_items if n in kpi_dict)
                        st.info(f"KPI 조회 완료: DB 미매칭 {len(no_db_items)}건 중 {matched_kpi}건 가격 확인")
                    elif not kpi_avail:
                        st.warning("KPI 시장가 조회를 하려면 사이드바에서 '한국물가정보 조회'를 ON 해주세요.")

                # 검토 DataFrame 구성
                review_df = build_review_df(result_df, kpi_dict, discount_rate_default)

                # data_editor — 할인율(%)과 검토근거만 편집 가능
                st.markdown("#### 📝 품목별 검토 결정표")
                REVIEW_OPTIONS = ["견적가 적용", "표준단가 적용", "할인율 적용", "시장가 적용"]

                display_cols = ["품명", "규격", "단위", "수량",
                                "견적단가", "DB존재", "표준단가", "표준단가_방향",
                                "KPI시장가", "할인율(%)", "할인율적용가",
                                "시스템추천", "검토근거"]

                edited = st.data_editor(
                    review_df[display_cols],
                    column_config={
                        "할인율(%)": st.column_config.NumberColumn(
                            "할인율(%)", min_value=0.0, max_value=50.0, step=0.1,
                            format="%.1f %%", help="품목별 협상 목표 할인율"
                        ),
                        "검토근거": st.column_config.SelectboxColumn(
                            "검토근거 ✏️", options=REVIEW_OPTIONS,
                            help="최종 적용 기준 선택"
                        ),
                        "견적단가":     st.column_config.NumberColumn("견적단가(원)", format="%d"),
                        "표준단가":     st.column_config.NumberColumn("표준단가(원)", format="%d"),
                        "KPI시장가":    st.column_config.NumberColumn("KPI시장가(원)", format="%d"),
                        "할인율적용가": st.column_config.NumberColumn("할인율적용가(원)", format="%d"),
                        "수량":         st.column_config.NumberColumn("수량", format="%.0f"),
                    },
                    disabled=["품명", "규격", "단위", "수량", "견적단가",
                              "DB존재", "표준단가", "표준단가_방향",
                              "KPI시장가", "할인율적용가", "시스템추천"],
                    hide_index=True,
                    use_container_width=True,
                    height=min(400, 40 + len(review_df) * 35),
                    key="review_editor"
                )

                # 편집 결과로 최종검토가 재계산
                final_df = review_df.copy()
                final_df["할인율(%)"] = edited["할인율(%)"].values
                final_df["검토근거"] = edited["검토근거"].values
                final_df["할인율적용가"] = (
                    final_df["견적단가"] * (1 - final_df["할인율(%)"] / 100)
                ).astype(int)
                final_df["최종검토가"] = final_df.apply(compute_final_price, axis=1)

                # 요약 메트릭
                st.markdown("#### 💰 검토 결과 요약")
                qty_col = final_df["수량"].fillna(1)
                총견적액 = (final_df["견적단가"] * qty_col).sum()
                총검토액 = (final_df["최종검토가"] * qty_col).sum()
                절감액   = 총견적액 - 총검토액
                절감률   = (절감액 / 총견적액 * 100) if 총견적액 > 0 else 0

                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("총 견적금액", f"{총견적액:,.0f}원")
                mc2.metric("검토 후 금액", f"{총검토액:,.0f}원",
                           delta=f"-{절감액:,.0f}원" if 절감액 > 0 else f"+{-절감액:,.0f}원",
                           delta_color="normal" if 절감액 > 0 else "inverse")
                mc3.metric("절감 가능 금액", f"{절감액:,.0f}원")
                mc4.metric("절감률", f"{절감률:.1f}%")

                # 최종 검토가 테이블
                st.markdown("#### ✅ 품목별 최종 검토가")
                final_show = final_df[["품명", "규격", "수량", "견적단가",
                                       "최종검토가", "검토근거", "시스템추천"]].copy()

                def color_review(row):
                    if row["검토근거"] != row["시스템추천"]:
                        return ["background-color: #FFF9C4"] * len(row)
                    if row["최종검토가"] < row["견적단가"]:
                        return ["background-color: #D4F5DC"] * len(row)
                    return [""] * len(row)

                st.dataframe(
                    final_show.style.apply(color_review, axis=1).format({
                        "견적단가":   "{:,.0f}",
                        "최종검토가": "{:,.0f}",
                        "수량":       lambda x: f"{x:.0f}" if pd.notna(x) else "-",
                    }),
                    use_container_width=True, height=min(400, 40 + len(final_df) * 35)
                )

                # 검토 확정 & DB 저장
                st.markdown("---")
                col_save, col_dl = st.columns([1, 2])
                with col_save:
                    if st.button("💾 검토 확정 & DB 저장", type="primary"):
                        final_df["DB존재 (내부)"] = review_df["DB존재"].map({"✅": True, "❌": False})
                        ok, msg = save_review_to_db(
                            final_df, uploaded.name,
                            str(Path(__file__).parent)
                        )
                        if ok:
                            st.success(msg)
                            st.cache_data.clear()
                            st.info("표준단가DB가 업데이트되었습니다. 사이드바 '캐시 초기화' 버튼을 눌러 DB를 갱신하세요.")
                        else:
                            st.error(msg)

                with col_dl:
                    # Excel 검토 리포트 다운로드
                    def make_review_excel(final_df, fname):
                        from openpyxl.styles import Font, PatternFill, Alignment
                        from openpyxl.utils import get_column_letter
                        wb2 = openpyxl.Workbook()
                        ws2 = wb2.active
                        ws2.title = "견적검토결과"
                        hfont = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
                        hfill = PatternFill('solid', fgColor='1F4E79')
                        center = Alignment(horizontal='center', vertical='center')

                        cols = ["품명", "규격", "단위", "수량",
                                "견적단가", "DB존재", "표준단가", "표준단가_방향",
                                "KPI시장가", "할인율(%)", "할인율적용가",
                                "시스템추천", "검토근거", "최종검토가"]
                        widths = [30, 25, 8, 8, 15, 8, 15, 18, 15, 10, 15, 16, 16, 15]

                        ws2.cell(1, 1, f"견적 검토 결과 — {fname}")
                        ws2.cell(1, 1).font = Font(name='맑은 고딕', bold=True, size=13, color='1F4E79')
                        ws2.merge_cells(f'A1:{get_column_letter(len(cols))}1')
                        ws2.row_dimensions[1].height = 24

                        for ci, (h, w) in enumerate(zip(cols, widths), 1):
                            c = ws2.cell(2, ci, h)
                            c.font = hfont; c.fill = hfill; c.alignment = center
                            ws2.column_dimensions[get_column_letter(ci)].width = w

                        num_cols_idx = [cols.index(c) + 1 for c in
                                        ["수량", "견적단가", "표준단가", "KPI시장가", "할인율적용가", "최종검토가"]
                                        if c in cols]
                        for ri, row in final_df[cols].iterrows():
                            for ci, val in enumerate(row, 1):
                                cell = ws2.cell(ri + 3, ci, val)
                                cell.font = Font(name='맑은 고딕', size=9)
                                cell.alignment = center
                            for ci in num_cols_idx:
                                ws2.cell(ri + 3, ci).number_format = '#,##0'
                            ws2.row_dimensions[ri + 3].height = 18

                        ws2.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(final_df)}"
                        ws2.freeze_panes = 'A3'
                        buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
                        return buf2

                    rev_buf = make_review_excel(final_df, uploaded.name)
                    st.download_button(
                        "📥 Excel 검토결과 다운로드",
                        data=rev_buf,
                        file_name=f"견적검토결과_{uploaded.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

# ── Tab 2: 표준단가 DB 조회 ────────────────────────────────────────────
with tab2:
    st.markdown("### 🗄️ 표준단가 DB")
    if db_df.empty:
        st.warning("DB를 로드할 수 없습니다.")
    else:
        search_kw = st.text_input("품목명 검색", placeholder="예) Air cylinder, 컨베어, BALL BUSH")
        filtered = db_df[db_df['품명'].str.contains(search_kw, case=False, na=False)] if search_kw else db_df

        c1, c2, c3 = st.columns(3)
        c1.metric("전체 품목", f"{len(db_df):,}개")
        c2.metric("검색 결과", f"{len(filtered):,}개")
        c3.metric("이상가 임계값", f"+{threshold_pct}%")

        st.dataframe(
            filtered.style.format({
                '단가_최저(원)': '{:,.0f}',
                '단가_평균(원)': '{:,.0f}',
                '단가_최고(원)': '{:,.0f}',
            }),
            use_container_width=True, height=600
        )

# ── Tab 3: 품목 단가 검색 ─────────────────────────────────────────────
with tab3:
    st.markdown("### 🔎 품목명으로 표준단가 직접 검색")
    if db_df.empty:
        st.warning("DB를 로드할 수 없습니다.")
    else:
        query_input = st.text_input("품목명 + 규격 입력", placeholder="예) Air cylinder CQ2A100 또는 고주파 가열기 50Kw")
        if query_input:
            with st.spinner("검색 중..."):
                vectorizer, matrix = build_search_index(db_df)
                results = find_similar(query_input, db_df, vectorizer, matrix, top_n=top_n)

            if results.empty:
                st.warning("유사 품목을 찾을 수 없습니다.")
            else:
                st.success(f"✅ {len(results)}개 유사 품목 발견")
                for _, r in results.iterrows():
                    sim_color = "#27AE60" if r['유사도'] > 0.5 else "#F39C12" if r['유사도'] > 0.2 else "#C0392B"
                    st.markdown(f"""
<div style="border:1px solid #DDD; border-radius:6px; padding:10px 14px; margin-bottom:8px; background:#FAFAFA;">
<b>{r['품명']}</b> &nbsp;|&nbsp; <span style="color:#666">{r.get('규격','')}</span> &nbsp;|&nbsp; 단위: {r.get('단위','')}<br>
단가 범위: <b>{int(r['단가_최저(원)']):,}원</b> ~ <b>{int(r['단가_평균(원)']):,}원</b> ~ <b>{int(r['단가_최고(원)']):,}원</b>
&nbsp;&nbsp;|&nbsp;&nbsp; 메이커: {r.get('주요메이커','')}
&nbsp;&nbsp;|&nbsp;&nbsp; 유사도: <b style="color:{sim_color}">{r['유사도']*100:.1f}%</b>
</div>
""", unsafe_allow_html=True)

            # KPI 시장가 조회
            if kpi_enabled and st.session_state.get("kpi_logged_in"):
                st.markdown("---")
                st.markdown("#### 🌐 한국물가정보(KPI) 시장가")
                with st.spinner("KPI 시장가 조회 중..."):
                    try:
                        kpi_items = kpi_search_subprocess(query_input.split()[0], max_results=10)
                    except Exception as e:
                        kpi_items = []
                        st.warning(f"KPI 조회 오류: {e}")

                if kpi_items:
                    st.success(f"✅ KPI {len(kpi_items)}건 조회")
                    kpi_df = pd.DataFrame(kpi_items)[["품명", "규격", "단위", "단가", "단가_최저", "단가_최고", "출처"]]
                    st.dataframe(
                        kpi_df.style.format({"단가": "{:,.0f}", "단가_최저": "{:,.0f}", "단가_최고": "{:,.0f}"}),
                        use_container_width=True
                    )
                else:
                    st.info("KPI에서 해당 품목을 찾을 수 없습니다.")
