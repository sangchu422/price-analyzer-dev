import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pathlib import Path
_here = Path(__file__).parent
with open(_here / "표준단가DB_집계.json", encoding='utf-8') as f:
    db_records = json.load(f)
with open(_here / "표준단가DB_전체원본.json", encoding='utf-8') as f:
    all_items = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "표준단가DB"

hfont = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
hfill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
right = Alignment(horizontal='right', vertical='center')
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 타이틀
ws.merge_cells('A1:L1')
ws['A1'] = f'협력사 견적 단가 AI 비교·분석 시스템 — 표준단가 DB'
ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14, color='1F3864')
ws['A1'].alignment = center
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:L2')
ws['A2'] = f'※ 11개 협력사 견적서 데이터 기반 구축 | 총 {len(db_records):,}개 품목 ({len(all_items):,}건 원본) | 이상가 임계값: +15%'
ws['A2'].font = Font(name='맑은 고딕', size=9, color='666666')
ws['A2'].alignment = center
ws.row_dimensions[2].height = 18

headers = ['표준품목ID', '품목분류', '품명', '규격', '단위',
           '단가_최저(원)', '단가_평균(원)', '단가_최고(원)',
           '데이터건수', '주요메이커', '최근견적일', '출처공급사']
widths   = [12, 10, 35, 30, 8, 15, 15, 15, 10, 15, 12, 25]

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=3, column=ci, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = center; cell.border = border
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[3].height = 22

for ri, rec in enumerate(db_records, 4):
    alt = (ri % 2 == 0)
    vals = [rec['표준품목ID'], rec['품목분류'], rec['품명'], rec['규격'], rec['단위'],
            rec['단가_최저'], rec['단가_평균'], rec['단가_최고'],
            rec['데이터건수'], rec['주요메이커'], rec['최근견적일'], rec['출처']]
    alns = [center, center, left, left, center, right, right, right, center, left, center, left]
    for ci, (v, a) in enumerate(zip(vals, alns), 1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = Font(name='맑은 고딕', size=9)
        cell.alignment = a; cell.border = border
        if alt: cell.fill = alt_fill
    for ci in [6, 7, 8]:
        ws.cell(row=ri, column=ci).number_format = '#,##0'
    ws.row_dimensions[ri].height = 18

ws.auto_filter.ref = f"A3:L{3 + len(db_records)}"
ws.freeze_panes = 'A4'

# 원본 데이터 시트
ws2 = wb.create_sheet("원본견적데이터")
h2 = ['ID', '공급사', '견적번호', '견적일', '공사명', '단위장비명', '품목분류',
      '품명', '규격', '단위', '수량', '단가(원)', '금액(원)', '메이커']
w2 = [6, 14, 20, 12, 25, 25, 10, 30, 25, 8, 8, 15, 15, 15]
for ci, (h, w) in enumerate(zip(h2, w2), 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = center; cell.border = border
    ws2.column_dimensions[get_column_letter(ci)].width = w

for ri, it in enumerate(all_items, 2):
    vals = [it['ID'], it['공급사'], it['견적번호'], it['견적일'], it['공사명'],
            it['단위장비명'], it['품목분류'], it['품명'], it['규격'],
            it['단위'], it['수량'], it['단가_원'], it['금액_원'], it['메이커']]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = Font(name='맑은 고딕', size=9)
        cell.border = border
        cell.alignment = center if ci in [1, 2, 3, 4, 7, 10, 11] else left
    for ci in [12, 13]:
        ws2.cell(row=ri, column=ci).number_format = '#,##0'
    ws2.row_dimensions[ri].height = 18

ws2.auto_filter.ref = f"A1:N{1 + len(all_items)}"
ws2.freeze_panes = 'A2'

out = str(_here / "표준단가DB.xlsx")
wb.save(out)
print(f"저장 완료: {out}")
print(f"  표준단가DB: {len(db_records):,}개 품목")
print(f"  원본견적데이터: {len(all_items):,}개 레코드")
