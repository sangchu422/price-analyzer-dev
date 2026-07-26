"""
xlsx / xls 파일 읽기 — 시트명: [[row값, ...], ...] 딕셔너리 반환
IRM 잠금 · 손상 파일은 예외를 발생시켜 호출자가 로그 처리.
"""
from pathlib import Path


def read_excel(path: Path) -> dict[str, list[list[str]]]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx(path)
    elif ext == ".xls":
        return _read_xls(path)
    raise ValueError(f"지원하지 않는 확장자: {ext}")


def _read_xlsx(path: Path) -> dict[str, list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(max_row=600, values_only=True):
            r = [_clean(v) for v in row]
            if any(r):
                rows.append(r)
        sheets[name] = rows
    return sheets


def _read_xls(path: Path) -> dict[str, list[list[str]]]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sheets = {}
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        rows = []
        for i in range(min(600, ws.nrows)):
            r = [_clean(ws.cell_value(i, j)) for j in range(ws.ncols)]
            if any(r):
                rows.append(r)
        sheets[name] = rows
    return sheets


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan") else s
